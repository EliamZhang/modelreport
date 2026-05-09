from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .metric_calculator import safe_divide


VARIABLE_NAME_COLUMNS = [
    "field",
    "fields",
    "variable",
    "variable_name",
    "metric",
    "feature",
    "feature_name",
]
KEY_COLUMNS = {"application_id", "user_id", "sample_datetime"}
MAX_EXCEL_ROWS = 1_048_576


def normalize_application_id(series: pd.Series) -> pd.Series:
    normalized = series.astype("string").str.strip()
    normalized = normalized.str.replace(r"\.0+$", "", regex=True)
    return normalized.mask(normalized.isin(["", "nan", "None", "<NA>"]), pd.NA)


def _find_column_case_insensitive(columns: list[str], candidates: list[str]) -> str | None:
    lower_to_original = {str(col).strip().lower(): col for col in columns}
    for candidate in candidates:
        if candidate in lower_to_original:
            return lower_to_original[candidate]
    return None


def load_variable_library(path: Path, logger) -> tuple[pd.DataFrame, list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"model_variable_library.csv not found: {path}")

    library_df = pd.read_csv(path, low_memory=False)
    if library_df.empty and not len(library_df.columns):
        raise ValueError(f"model_variable_library.csv has no columns: {path}")

    columns = [str(col) for col in library_df.columns]
    field_col = _find_column_case_insensitive(columns, VARIABLE_NAME_COLUMNS)
    if field_col:
        variables = (
            library_df[field_col]
            .dropna()
            .astype("string")
            .str.strip()
            .loc[lambda s: s.ne("")]
            .drop_duplicates()
            .tolist()
        )
    else:
        variables = [col for col in columns if col not in KEY_COLUMNS]
        logger.warning(
            "No variable-name column found in model_variable_library.csv; "
            "treating wide-format non-key columns as variables."
        )

    if not variables:
        raise ValueError(f"No variables found in model_variable_library.csv: {path}")

    return library_df, variables


def merge_variable_library_data(
    df: pd.DataFrame,
    library_df: pd.DataFrame,
    variables: list[str],
    primary_key: str,
    logger,
) -> pd.DataFrame:
    if primary_key not in df.columns:
        raise ValueError(f"Main analysis data missing primary key: {primary_key}")

    out = df.copy()
    out[primary_key] = normalize_application_id(out[primary_key])

    if primary_key not in library_df.columns:
        logger.warning(
            f"model_variable_library.csv does not contain {primary_key}; "
            "only variables already present in main analysis data can be analyzed."
        )
        return out

    merge_fields = [field for field in variables if field in library_df.columns and field not in out.columns]
    if not merge_fields:
        return out

    right = library_df[[primary_key] + merge_fields].copy()
    right[primary_key] = normalize_application_id(right[primary_key])
    duplicate_rows = int(right.duplicated(subset=[primary_key], keep=False).sum())
    if duplicate_rows:
        logger.warning(
            "model_variable_library.csv has duplicate application_id values; keeping first: "
            f"duplicate_rows={duplicate_rows:,}"
        )
        right = right.drop_duplicates(subset=[primary_key], keep="first")

    before_rows = len(out)
    out = out.merge(right, on=primary_key, how="left")
    if len(out) != before_rows:
        raise ValueError(f"Variable library merge changed row count: before={before_rows}, after={len(out)}")
    logger.info(f"Merged model variable library fields: added_field_cnt={len(merge_fields):,}")
    return out


def _existing_and_missing_fields(df: pd.DataFrame, variables: list[str]) -> tuple[list[str], list[str]]:
    existing = [field for field in variables if field in df.columns]
    missing = [field for field in variables if field not in df.columns]
    return existing, missing


def _to_numeric(df: pd.DataFrame, variable: str) -> pd.Series:
    return pd.to_numeric(df[variable], errors="coerce")


def _sort_values(values: list[Any]) -> list[Any]:
    def sort_key(value: Any) -> tuple[int, float, str]:
        if value is None or pd.isna(value):
            return (2, 0.0, "")
        try:
            return (0, float(value), str(value))
        except (TypeError, ValueError):
            return (1, 0.0, str(value))

    return sorted(values, key=sort_key)


def build_mean_by_primary_bin(
    df: pd.DataFrame,
    variables: list[str],
    primary_bin_col: str,
) -> pd.DataFrame:
    return _build_mean_by_bin(df, variables, primary_bin_col)


def build_mean_by_comparison_bin(
    df: pd.DataFrame,
    variables: list[str],
    comparison_bin_col: str,
) -> pd.DataFrame:
    return _build_mean_by_bin(df, variables, comparison_bin_col)


def _build_mean_by_bin(df: pd.DataFrame, variables: list[str], bin_col: str) -> pd.DataFrame:
    if bin_col not in df.columns:
        raise ValueError(f"Main analysis data missing {bin_col}")

    records: list[dict[str, Any]] = []
    for bin_value in _sort_values(df[bin_col].drop_duplicates().tolist()):
        group = df[df[bin_col].astype(str).eq(str(bin_value))]
        total_cnt = len(group)
        for variable in variables:
            numeric = _to_numeric(group, variable)
            non_null_cnt = int(numeric.notna().sum())
            records.append(
                {
                    bin_col: bin_value,
                    "variable_name": variable,
                    "avg_value": float(numeric.mean()) if non_null_cnt else None,
                    "non_null_cnt": non_null_cnt,
                    "total_cnt": total_cnt,
                    "missing_pct": safe_divide(total_cnt - non_null_cnt, total_cnt),
                }
            )
    return pd.DataFrame(records)


def build_summary_by_variable(df: pd.DataFrame, variables: list[str]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    total_cnt = len(df)
    for variable in variables:
        numeric = _to_numeric(df, variable)
        non_null_cnt = int(numeric.notna().sum())
        missing_cnt = total_cnt - non_null_cnt
        records.append(
            {
                "variable_name": variable,
                "avg_value": float(numeric.mean()) if non_null_cnt else None,
                "non_null_cnt": non_null_cnt,
                "total_cnt": total_cnt,
                "missing_cnt": missing_cnt,
                "missing_pct": safe_divide(missing_cnt, total_cnt),
                "min_value": float(numeric.min()) if non_null_cnt else None,
                "max_value": float(numeric.max()) if non_null_cnt else None,
            }
        )
    return pd.DataFrame(records)


def build_cross_bin_mean_pivot(
    df: pd.DataFrame,
    variable: str,
    primary_bin_col: str,
    comparison_bin_col: str,
) -> pd.DataFrame:
    missing_cols = [col for col in [primary_bin_col, comparison_bin_col] if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Main analysis data missing columns: {missing_cols}")

    primary_values = _sort_values(df[primary_bin_col].drop_duplicates().tolist())
    comparison_values = _sort_values(df[comparison_bin_col].drop_duplicates().tolist())
    numeric = _to_numeric(df, variable)

    records: list[dict[str, Any]] = []
    for primary_value in primary_values:
        row_mask = df[primary_bin_col].astype(str).eq(str(primary_value))
        record: dict[str, Any] = {primary_bin_col: primary_value}
        for comparison_value in comparison_values:
            cell_mask = row_mask & df[comparison_bin_col].astype(str).eq(str(comparison_value))
            record[str(comparison_value)] = _mean_or_none(numeric[cell_mask])
        record["Total"] = _mean_or_none(numeric[row_mask])
        records.append(record)

    total_record: dict[str, Any] = {primary_bin_col: "Total"}
    for comparison_value in comparison_values:
        column_mask = df[comparison_bin_col].astype(str).eq(str(comparison_value))
        total_record[str(comparison_value)] = _mean_or_none(numeric[column_mask])
    total_record["Total"] = _mean_or_none(numeric)
    records.append(total_record)

    return pd.DataFrame(records, columns=[primary_bin_col] + [str(v) for v in comparison_values] + ["Total"])


def _mean_or_none(series: pd.Series) -> float | None:
    non_null = series.dropna()
    if non_null.empty:
        return None
    return float(non_null.mean())


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1) -> tuple[int, int]:
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=1 + col_offset, value=None if pd.isna(value) else value)
    return len(rows), len(df.columns)


def _style_table(ws, start_row: int, row_count: int, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    end_row = start_row + row_count - 1

    for col in range(1, col_count + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in range(start_row + 1, end_row + 1):
        is_total_row = ws.cell(row=row, column=1).value == "Total"
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col > 1:
                cell.number_format = "0.0000"
            if is_total_row or ws.cell(row=start_row, column=col).value == "Total":
                cell.fill = total_fill


def _style_flat_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = headers[cell.column - 1]
            cell.border = border
            if header in {"avg_value", "min_value", "max_value"}:
                cell.number_format = "0.0000"
            elif header in {"missing_pct"}:
                cell.number_format = "0.00%"
            elif header in {"non_null_cnt", "total_cnt", "missing_cnt"}:
                cell.number_format = "#,##0"


def _auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for item in cell:
                if item.value is not None:
                    max_len = max(max_len, len(str(item.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _safe_sheet_name(base_name: str, part: int | None = None) -> str:
    suffix = "" if part is None else f"_{part}"
    return f"{base_name[:31 - len(suffix)]}{suffix}"


def _write_split_dataframe(wb: Workbook, base_sheet_name: str, df: pd.DataFrame) -> None:
    max_data_rows = MAX_EXCEL_ROWS - 1
    if len(df) <= max_data_rows:
        ws = wb.create_sheet(_safe_sheet_name(base_sheet_name))
        _write_dataframe(ws, df)
        _style_flat_sheet(ws)
        _auto_width(ws)
        return

    start = 0
    part = 1
    while start < len(df):
        chunk = df.iloc[start : start + max_data_rows]
        ws = wb.create_sheet(_safe_sheet_name(base_sheet_name, part))
        _write_dataframe(ws, chunk)
        _style_flat_sheet(ws)
        _auto_width(ws)
        start += max_data_rows
        part += 1


def write_cross_bin_mean_pivot_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    variables: list[str],
    primary_bin_col: str,
    comparison_bin_col: str,
) -> None:
    ws = wb.create_sheet("cross_bin_distribution")
    ws.freeze_panes = "A2"
    current_row = 1
    for variable in variables:
        ws.cell(row=current_row, column=1, value=f"{variable}_avg").font = Font(bold=True, size=12)
        current_row += 1
        pivot = build_cross_bin_mean_pivot(df, variable, primary_bin_col, comparison_bin_col)
        row_count, col_count = _write_dataframe(ws, pivot, current_row)
        _style_table(ws, current_row, row_count, col_count)
        current_row += row_count + 2
    _auto_width(ws)


def export_discrete_distribution_excel(
    output_path: Path,
    analysis_df: pd.DataFrame,
    variables: list[str],
    primary_distribution: pd.DataFrame,
    comparison_distribution: pd.DataFrame,
    summary_by_variable: pd.DataFrame,
    primary_bin_col: str,
    comparison_bin_col: str,
    logger,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    _write_split_dataframe(wb, "primary_bin_distribution", primary_distribution)
    _write_split_dataframe(wb, "comparison_bin_distribution", comparison_distribution)
    write_cross_bin_mean_pivot_sheet(wb, analysis_df, variables, primary_bin_col, comparison_bin_col)
    _write_split_dataframe(wb, "summary_by_variable", summary_by_variable)

    wb.save(output_path)
    logger.info(f"Wrote variable mean distribution Excel: {output_path}")


def build_and_export_discrete_variable_distribution(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    logger,
    primary_bin_col: str = "primary_model_score_bin",
    comparison_bin_col: str = "comparison_model_score_bin",
) -> Path:
    project_root = Path(cfg["project"]["root_dir"])
    source_path = project_root / "INPUT" / "model_variable_library.csv"
    output_dir = Path(cfg["project"]["output_dir"])
    output_path = output_dir / "model_bin_feature_mean_profile.xlsx"
    primary_key = cfg["keys"]["primary_key"]

    output_dir.mkdir(parents=True, exist_ok=True)
    for old_workbook in output_dir.glob("*.xlsx"):
        if old_workbook.name != output_path.name:
            old_workbook.unlink()
    legacy_output_dir = project_root / "OUTPUT"
    legacy_output_path = legacy_output_dir / output_path.name
    if legacy_output_path.exists():
        legacy_output_path.unlink()

    library_df, variables = load_variable_library(source_path, logger)
    analysis_df = merge_variable_library_data(df, library_df, variables, primary_key, logger)
    if primary_bin_col not in analysis_df.columns:
        raise ValueError(f"Main analysis data missing {primary_bin_col}")
    if comparison_bin_col not in analysis_df.columns:
        raise ValueError(f"Main analysis data missing {comparison_bin_col}")

    existing_fields, missing_field_names = _existing_and_missing_fields(analysis_df, variables)
    if not existing_fields:
        raise ValueError("No variables from model_variable_library.csv matched main analysis data columns.")

    summary = build_summary_by_variable(analysis_df, existing_fields)
    all_null_cnt = int(summary["non_null_cnt"].eq(0).sum())
    for variable in summary.loc[summary["non_null_cnt"].eq(0), "variable_name"]:
        logger.warning(f"Variable has no numeric values after pd.to_numeric conversion: {variable}")

    logger.info(f"Variable list read count: {len(variables):,}")
    logger.info(f"Variables matched count: {len(existing_fields):,}")
    logger.info(f"Variables missing count: {len(missing_field_names):,}")
    logger.info(f"All-null or non-numeric variable count: {all_null_cnt:,}")
    logger.info(f"Variable mean distribution Excel output path: {output_path}")

    primary_distribution = build_mean_by_primary_bin(analysis_df, existing_fields, primary_bin_col)
    comparison_distribution = build_mean_by_comparison_bin(analysis_df, existing_fields, comparison_bin_col)

    export_discrete_distribution_excel(
        output_path=output_path,
        analysis_df=analysis_df,
        variables=existing_fields,
        primary_distribution=primary_distribution,
        comparison_distribution=comparison_distribution,
        summary_by_variable=summary,
        primary_bin_col=primary_bin_col,
        comparison_bin_col=comparison_bin_col,
        logger=logger,
    )
    return output_path
