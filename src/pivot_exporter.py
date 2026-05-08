from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_METRICS = [
    "sample_cnt",
    "sample_pct",
    "duedate_3m_30_valid_cnt",
    "duedate_3m_30_bad_cnt",
    "duedate_3m_30_good_cnt",
    "duedate_3m_30_bad_rate",
    "duedate_1m_5_valid_cnt",
    "duedate_1m_5_bad_cnt",
    "duedate_1m_5_good_cnt",
    "duedate_1m_5_bad_rate",
    "duedate_3m_30_amount_overdue_amount",
    "duedate_3m_30_amount_principal_amount",
    "duedate_3m_30_amount_overdue_rate",
    "duedate_1m_5_amount_overdue_amount",
    "duedate_1m_5_amount_principal_amount",
    "duedate_1m_5_amount_overdue_rate",
    "avg_PTI",
    "avg_requested_loan_amount",
    "avg_total_amount",
    "avg_gross_surplus",
    "avg_net_surplus",
    "avg_total_income",
    "avg_bank_txn_industry_primary_industry_amt",
    "avg_bank_txn_industry_pay_interval_median",
    "avg_bank_txn_industry_pay_interval_std",
    "avg_bank_txn_industry_primary_employer_ratio",
    "avg_bank_txn_industry_employer_cnt",
    "apply_cnt",
    "completed_application_cnt",
    "approved_application_cnt",
    "auto_approved_application_cnt",
    "manual_approved_application_cnt",
    "deal_sample_cnt",
    "completion_rate",
    "approval_rate",
    "auto_approval_rate",
    "manual_approval_rate",
    "auto_approval_share",
    "manual_approval_share",
    "deal_rate",
]

MISSING_LABEL = "Missing"
TOTAL_LABEL = "Total"


def _read_input_table(input_file: str | Path) -> pd.DataFrame:
    path = Path(input_file)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    return pd.read_csv(path)


def _ensure_output_path(output_excel: str | Path) -> Path:
    path = Path(output_excel)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _is_missing(value: Any) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _normalize_dim_value(value: Any) -> Any:
    if _is_missing(value):
        return MISSING_LABEL
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _sort_dim_values(values: pd.Series) -> list[Any]:
    normalized = [_normalize_dim_value(v) for v in values.drop_duplicates().tolist()]

    def sort_key(value: Any) -> tuple[int, float | str]:
        if value == MISSING_LABEL:
            return (2, "")
        numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        if not pd.isna(numeric):
            return (0, float(numeric))
        return (1, str(value))

    return sorted(normalized, key=sort_key)


def _is_rate_metric(metric: str) -> bool:
    metric_lower = metric.lower()
    return metric_lower.endswith("_rate") or metric_lower.endswith("_pct") or metric_lower.endswith("_share")


def _is_count_metric(metric: str) -> bool:
    metric_lower = metric.lower()
    count_tokens = ("_cnt", "_amount", "_principal_amount", "_overdue_amount")
    return metric_lower.endswith(count_tokens) or metric_lower in {"sample_cnt", "apply_cnt"}


def _is_average_metric(metric: str) -> bool:
    return metric.lower().startswith("avg_")


def _infer_ratio_columns(metric: str, df: pd.DataFrame) -> tuple[str, str] | None:
    if metric == "sample_pct" and "sample_cnt" in df.columns:
        return ("sample_cnt", "sample_cnt")

    conversion_ratios = {
        "completion_rate": ("completed_application_cnt", "apply_cnt"),
        "approval_rate": ("approved_application_cnt", "completed_application_cnt"),
        "auto_approval_rate": ("auto_approved_application_cnt", "completed_application_cnt"),
        "manual_approval_rate": ("manual_approved_application_cnt", "completed_application_cnt"),
        "auto_approval_share": ("auto_approved_application_cnt", "approved_application_cnt"),
        "manual_approval_share": ("manual_approved_application_cnt", "approved_application_cnt"),
        "deal_rate": ("deal_sample_cnt", "approved_application_cnt"),
    }
    if metric in conversion_ratios:
        numerator, denominator = conversion_ratios[metric]
        if numerator in df.columns and denominator in df.columns:
            return (numerator, denominator)

    if metric.endswith("_bad_rate"):
        prefix = metric[: -len("_bad_rate")]
        numerator = f"{prefix}_bad_cnt"
        denominator = f"{prefix}_valid_cnt"
        if numerator in df.columns and denominator in df.columns:
            return (numerator, denominator)

    if metric.endswith("_overdue_rate"):
        prefix = metric[: -len("_overdue_rate")]
        numerator = f"{prefix}_overdue_amount"
        denominator = f"{prefix}_principal_amount"
        if numerator in df.columns and denominator in df.columns:
            return (numerator, denominator)

    return None


def _safe_divide(numerator: float, denominator: float) -> float | None:
    if pd.isna(denominator) or denominator == 0:
        return None
    if pd.isna(numerator):
        return None
    return float(numerator) / float(denominator)


def _sum_group(df: pd.DataFrame, rows: list[Any], cols: list[Any], value_col: str) -> pd.DataFrame:
    work = df[["__row_dim", "__col_dim", value_col]].copy()
    work[value_col] = pd.to_numeric(work[value_col], errors="coerce")
    grouped = work.groupby(["__row_dim", "__col_dim"], dropna=False)[value_col].sum(min_count=1)
    pivot = grouped.unstack("__col_dim").reindex(index=rows, columns=cols)
    return pivot


def _build_sum_pivot(df: pd.DataFrame, rows: list[Any], cols: list[Any], metric: str) -> pd.DataFrame:
    pivot = _sum_group(df, rows, cols, metric)
    pivot[TOTAL_LABEL] = pivot.sum(axis=1, min_count=1)

    total_row = pivot[cols].sum(axis=0, min_count=1)
    total_row[TOTAL_LABEL] = total_row.sum()
    pivot.loc[TOTAL_LABEL] = total_row
    return pivot


def _build_ratio_pivot(
    df: pd.DataFrame,
    rows: list[Any],
    cols: list[Any],
    metric: str,
    numerator_col: str,
    denominator_col: str,
) -> pd.DataFrame:
    numerator = _sum_group(df, rows, cols, numerator_col)
    denominator = _sum_group(df, rows, cols, denominator_col)

    if metric == "sample_pct":
        grand_denominator = denominator.to_numpy(dtype=float).sum()
        pivot = numerator.map(lambda value: _safe_divide(value, grand_denominator))
        row_numerator = numerator.sum(axis=1, min_count=1)
        col_numerator = numerator.sum(axis=0, min_count=1)
        pivot[TOTAL_LABEL] = row_numerator.map(lambda value: _safe_divide(value, grand_denominator))
        total_row = col_numerator.map(lambda value: _safe_divide(value, grand_denominator))
        total_row[TOTAL_LABEL] = _safe_divide(row_numerator.sum(), grand_denominator)
        pivot.loc[TOTAL_LABEL] = total_row
        return pivot

    pivot = numerator / denominator.where(denominator != 0)

    row_numerator = numerator.sum(axis=1, min_count=1)
    row_denominator = denominator.sum(axis=1, min_count=1)
    pivot[TOTAL_LABEL] = [
        _safe_divide(num, den) for num, den in zip(row_numerator.tolist(), row_denominator.tolist())
    ]

    col_numerator = numerator.sum(axis=0, min_count=1)
    col_denominator = denominator.sum(axis=0, min_count=1)
    total_row = pd.Series(
        [_safe_divide(num, den) for num, den in zip(col_numerator.tolist(), col_denominator.tolist())],
        index=cols,
    )
    total_row[TOTAL_LABEL] = _safe_divide(col_numerator.sum(), col_denominator.sum())
    pivot.loc[TOTAL_LABEL] = total_row
    return pivot


def _build_weighted_average_pivot(
    df: pd.DataFrame,
    rows: list[Any],
    cols: list[Any],
    metric: str,
    weight_col: str = "sample_cnt",
) -> pd.DataFrame:
    if weight_col not in df.columns:
        return _build_raw_rate_pivot(df, rows, cols, metric)

    work = df[["__row_dim", "__col_dim", metric, weight_col]].copy()
    work[metric] = pd.to_numeric(work[metric], errors="coerce")
    work[weight_col] = pd.to_numeric(work[weight_col], errors="coerce").fillna(0)
    work["__weighted_value"] = work[metric] * work[weight_col]
    work.loc[work[metric].isna(), ["__weighted_value", weight_col]] = 0

    weighted_sum = _sum_group(work, rows, cols, "__weighted_value")
    weight_sum = _sum_group(work, rows, cols, weight_col)
    pivot = weighted_sum / weight_sum.where(weight_sum != 0)

    row_weighted = weighted_sum.sum(axis=1, min_count=1)
    row_weight = weight_sum.sum(axis=1, min_count=1)
    pivot[TOTAL_LABEL] = [
        _safe_divide(num, den) for num, den in zip(row_weighted.tolist(), row_weight.tolist())
    ]

    col_weighted = weighted_sum.sum(axis=0, min_count=1)
    col_weight = weight_sum.sum(axis=0, min_count=1)
    total_row = pd.Series(
        [_safe_divide(num, den) for num, den in zip(col_weighted.tolist(), col_weight.tolist())],
        index=cols,
    )
    total_row[TOTAL_LABEL] = _safe_divide(col_weighted.sum(), col_weight.sum())
    pivot.loc[TOTAL_LABEL] = total_row
    return pivot


def _build_raw_rate_pivot(df: pd.DataFrame, rows: list[Any], cols: list[Any], metric: str) -> pd.DataFrame:
    # Unrecognized rate/pct metrics need manual confirmation because their numerator and denominator
    # cannot be inferred safely from the metric name and available columns.
    work = df[["__row_dim", "__col_dim", metric]].copy()
    work[metric] = pd.to_numeric(work[metric], errors="coerce")
    grouped = work.groupby(["__row_dim", "__col_dim"], dropna=False)[metric].first()
    pivot = grouped.unstack("__col_dim").reindex(index=rows, columns=cols)
    pivot[TOTAL_LABEL] = None
    pivot.loc[TOTAL_LABEL] = [None] * len(pivot.columns)
    return pivot


def _build_metric_pivot(df: pd.DataFrame, rows: list[Any], cols: list[Any], metric: str) -> pd.DataFrame:
    ratio_cols = _infer_ratio_columns(metric, df)
    if ratio_cols is not None:
        return _build_ratio_pivot(df, rows, cols, metric, ratio_cols[0], ratio_cols[1])
    if _is_average_metric(metric):
        return _build_weighted_average_pivot(df, rows, cols, metric)
    if _is_rate_metric(metric):
        return _build_raw_rate_pivot(df, rows, cols, metric)
    if _is_count_metric(metric):
        return _build_sum_pivot(df, rows, cols, metric)
    return _build_sum_pivot(df, rows, cols, metric)


def _format_sheet(writer: pd.ExcelWriter, sheet_name: str, metric: str) -> None:
    ws = writer.book[sheet_name]
    ws.freeze_panes = "B2"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill

    max_row = ws.max_row
    max_col = ws.max_column
    for cell in ws[max_row]:
        cell.font = bold_font
        cell.fill = total_fill

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=max_col, max_col=max_col):
        row[0].font = bold_font
        row[0].fill = total_fill

    number_format = "0.00%" if _is_rate_metric(metric) else "#,##0.00" if _is_average_metric(metric) else "#,##0"
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_col):
        for cell in row:
            cell.number_format = number_format

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in ws[letter])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 32)


def _sheet_name(metric: str, used_names: set[str]) -> str:
    invalid_chars = set('[]:*?/\\')
    safe = "".join(ch if ch not in invalid_chars else "_" for ch in metric).strip() or "metric"
    base = safe[:31]
    candidate = base
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def generate_pivot_excel(
    input_file: str,
    output_excel: str,
    row_dim: str = "primary_model_score_bin",
    col_dim: str = "comparison_model_score_bin",
    metrics: list[str] | None = None,
) -> Path:
    """Generate one Excel pivot sheet per metric with row, column, and grand totals."""
    df = _read_input_table(input_file)
    required_cols = {row_dim, col_dim}
    missing_dims = sorted(required_cols - set(df.columns))
    if missing_dims:
        raise ValueError(f"Pivot dimensions missing from input_file={input_file}: {missing_dims}")

    metric_list = list(metrics or DEFAULT_METRICS)
    missing_metrics = [metric for metric in metric_list if metric not in df.columns]
    if missing_metrics:
        raise ValueError(f"Pivot metrics missing from input_file={input_file}: {missing_metrics}")

    work = df.copy()
    work["__row_dim"] = work[row_dim].map(_normalize_dim_value)
    work["__col_dim"] = work[col_dim].map(_normalize_dim_value)

    rows = _sort_dim_values(work["__row_dim"])
    cols = _sort_dim_values(work["__col_dim"])
    output_path = _ensure_output_path(output_excel)

    used_names: set[str] = set()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for metric in metric_list:
            pivot = _build_metric_pivot(work, rows, cols, metric)
            out = pivot.reset_index().rename(columns={"__row_dim": row_dim, "index": row_dim})
            sheet_name = _sheet_name(metric, used_names)
            out.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer, sheet_name, metric)

    return output_path
