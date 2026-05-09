from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .metric_calculator import (
    calculate_group_metrics,
    calculate_user_profile_category_distribution,
    calculate_user_profile_numeric_metrics,
)


METRIC_GROUPS: dict[str, list[str]] = {
    "risk_perf": [
        "duedate_3m_30_valid_cnt",
        "duedate_3m_30_bad_cnt",
        "duedate_3m_30_good_cnt",
        "duedate_3m_30_bad_rate",
        "duedate_1m_5_valid_cnt",
        "duedate_1m_5_bad_cnt",
        "duedate_1m_5_good_cnt",
        "duedate_1m_5_bad_rate",
    ],
    "amount_risk": [
        "duedate_3m_30_amount_overdue_amount",
        "duedate_3m_30_amount_principal_amount",
        "duedate_3m_30_amount_overdue_rate",
        "duedate_1m_5_amount_overdue_amount",
        "duedate_1m_5_amount_principal_amount",
        "duedate_1m_5_amount_overdue_rate",
    ],
    "profile": [
        "avg_PTI",
        "avg_requested_loan_amount",
        "avg_total_amount",
        "avg_gross_surplus",
        "avg_net_surplus",
        "avg_total_income",
    ],
    "conversion": [
        "apply_cnt",
        "completed_application_cnt",
        "approved_application_cnt",
        "auto_approved_application_cnt",
        "manual_approved_application_cnt",
        "deal_sample_cnt",
        "completion_rate",
        "approval_rate",
        "auto_approval_rate",
        "manual_approval_rate",
        "auto_approval_share",
        "manual_approval_share",
        "deal_rate",
    ],
}

RATE_DENOMINATORS = {
    "duedate_3m_30_bad_rate": "duedate_3m_30_valid_cnt",
    "duedate_1m_5_bad_rate": "duedate_1m_5_valid_cnt",
    "duedate_3m_30_amount_overdue_rate": "duedate_3m_30_amount_principal_amount",
    "duedate_1m_5_amount_overdue_rate": "duedate_1m_5_amount_principal_amount",
    "completion_rate": "apply_cnt",
    "approval_rate": "completed_application_cnt",
    "auto_approval_rate": "completed_application_cnt",
    "manual_approval_rate": "completed_application_cnt",
    "auto_approval_share": "approved_application_cnt",
    "manual_approval_share": "approved_application_cnt",
    "deal_rate": "approved_application_cnt",
}

COUNT_METRICS = {
    "sample_cnt",
    "duedate_3m_30_valid_cnt",
    "duedate_3m_30_bad_cnt",
    "duedate_3m_30_good_cnt",
    "duedate_1m_5_valid_cnt",
    "duedate_1m_5_bad_cnt",
    "duedate_1m_5_good_cnt",
    "apply_cnt",
    "completed_application_cnt",
    "approved_application_cnt",
    "auto_approved_application_cnt",
    "manual_approved_application_cnt",
    "deal_sample_cnt",
}

AMOUNT_METRICS = {
    "duedate_3m_30_amount_overdue_amount",
    "duedate_3m_30_amount_principal_amount",
    "duedate_1m_5_amount_overdue_amount",
    "duedate_1m_5_amount_principal_amount",
}

FOUR_DECIMAL_METRICS = {"avg_PTI"}
USER_PROFILE_CATEGORY_ORDER = [
    "basic_profile",
    "occupation_profile",
]
USER_PROFILE_DISTRIBUTION_SHEET_ORDER = [
    "geo_profile_distribution",
    "family_profile_distribution",
    "attributed_category_distribution",
    "occupation_profile_distribution",
]
USER_PROFILE_BIN_ORDER = [1, 2, 3]


def _is_missing(value: Any) -> bool:
    return value is None or pd.isna(value)


def _label_key(value: Any) -> str:
    if _is_missing(value):
        return "__NA__"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _label_display(value: Any) -> str:
    if _is_missing(value):
        return "NA"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sort_values(values: list[Any]) -> list[Any]:
    def sort_key(value: Any) -> tuple[int, float, str]:
        if _is_missing(value):
            return (2, 0.0, "")
        try:
            return (0, float(value), str(value))
        except (TypeError, ValueError):
            return (1, 0.0, str(value))

    unique: dict[str, Any] = {}
    for value in values:
        unique.setdefault(_label_key(value), value)
    return sorted(unique.values(), key=sort_key)


def _records_by_key(metrics: pd.DataFrame, group_cols: list[str]) -> dict[tuple[str, ...], pd.Series]:
    records: dict[tuple[str, ...], pd.Series] = {}
    for _, row in metrics.iterrows():
        records[tuple(_label_key(row[col]) for col in group_cols)] = row
    return records


def _format_number(value: Any, metric: str) -> str:
    if _is_missing(value):
        return ""
    value = float(value)
    if metric in COUNT_METRICS:
        return f"{int(round(value)):,}"
    if metric in AMOUNT_METRICS:
        return f"{value:,.2f}"
    if metric in FOUR_DECIMAL_METRICS:
        return f"{value:.4f}"
    if metric.startswith("avg_"):
        return f"{value:,.2f}"
    return f"{value:,.4f}"


def _format_cell(record: pd.Series | None, metric: str) -> str:
    if record is None or metric not in record.index:
        return ""
    value = record[metric]
    if metric in RATE_DENOMINATORS:
        return "" if _is_missing(value) else f"{float(value):.2%}"
    return _format_number(value, metric)


def _metric_cell_format(metric: str) -> str:
    if metric.endswith("_pct") or metric.endswith("_rate") or metric.endswith("_share"):
        return "percent"
    if metric == "sample_cnt" or metric.endswith("_cnt"):
        return "integer"
    if metric.startswith("avg_"):
        return "decimal_2"
    return "value"


def _format_user_profile_value(value: Any, metric: str) -> str:
    if _is_missing(value) or value == "":
        return ""
    cell_format = _metric_cell_format(metric)
    if cell_format == "percent":
        return f"{float(value):.2%}"
    if cell_format == "integer":
        return f"{int(round(float(value))):,}"
    if cell_format == "decimal_2":
        return f"{float(value):,.2f}"
    return str(value)


def _apply_metric_number_format(ws, start_row: int, row_count: int, headers: list[str]) -> None:
    formats = {
        "percent": "0.00%",
        "integer": "#,##0",
        "decimal_2": "#,##0.00",
    }
    for col_idx, header in enumerate(headers, start=1):
        number_format = formats.get(_metric_cell_format(str(header)))
        if not number_format:
            continue
        for row_idx in range(start_row + 1, start_row + row_count):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format


def _user_profile_numeric_metric_specs(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []

    for metric_cfg in metrics:
        source_field = metric_cfg.get("source_field", "")
        metric = metric_cfg.get("output_field", f"avg_{source_field}")
        specs.append(
            {
                "category": metric_cfg.get("category", "user_profile"),
                "metric_type": "numeric_cross",
                "field_name": source_field,
                "metric_name": metric,
                "cell_format": _metric_cell_format(metric),
                "description": (
                    f"Mean value of {source_field} by primary_model_score_bin "
                    "and comparison_model_score_bin."
                ),
            }
        )
    return specs


def _user_profile_category_metric_specs(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for metric_cfg in metrics:
        source_field = metric_cfg.get("source_field", "")
        sheet = metric_cfg.get("sheet", "category_distribution")
        top_n = int(metric_cfg.get("top_n", 3))
        specs.extend(
            [
                {
                    "category": sheet,
                    "metric_type": "category_distribution",
                    "field_name": source_field,
                    "metric_name": f"{source_field}_cnt",
                    "cell_format": "integer",
                    "description": (
                        f"Count distribution of {source_field} by primary_model_score_bin; "
                        f"keeps global Top{top_n} values and combines the rest into Others."
                    ),
                },
                {
                    "category": sheet,
                    "metric_type": "category_distribution",
                    "field_name": source_field,
                    "metric_name": f"{source_field}_pct",
                    "cell_format": "percent",
                    "description": (
                        f"Share distribution of {source_field} by primary_model_score_bin; "
                        "denominator is the sample count of the current primary bin."
                    ),
                },
            ]
        )
    return specs


def _user_profile_metric_guide(
    numeric_metrics: list[dict[str, Any]],
    category_metrics: list[dict[str, Any]],
) -> pd.DataFrame:
    columns = ["category", "metric_type", "field_name", "metric_name", "cell_format", "description"]
    records = _user_profile_numeric_metric_specs(numeric_metrics) + _user_profile_category_metric_specs(category_metrics)
    return pd.DataFrame(records, columns=columns)


def _user_profile_numeric_metric_groups(metrics: list[dict[str, Any]]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {}
    for metric_cfg in metrics:
        category = metric_cfg.get("category", "user_profile")
        metric = metric_cfg.get("output_field", f"avg_{metric_cfg.get('source_field', '')}")
        grouped.setdefault(category, []).append(metric)

    ordered: dict[str, list[str]] = {}
    for category in USER_PROFILE_CATEGORY_ORDER:
        if category in grouped:
            ordered[category] = grouped.pop(category)
    ordered.update(grouped)
    return ordered


def _build_metric_pivot(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    metric: str,
) -> pd.DataFrame:
    detail = calculate_group_metrics(df, [row_bin, column_bin], cfg)
    row_totals = calculate_group_metrics(df, [row_bin], cfg)
    column_totals = calculate_group_metrics(df, [column_bin], cfg)
    grand_total = calculate_group_metrics(df, [], cfg)

    row_values = _sort_values(list(detail[row_bin]) + list(row_totals[row_bin]))
    column_values = _sort_values(list(detail[column_bin]) + list(column_totals[column_bin]))

    detail_records = _records_by_key(detail, [row_bin, column_bin])
    row_total_records = _records_by_key(row_totals, [row_bin])
    column_total_records = _records_by_key(column_totals, [column_bin])
    grand_record = grand_total.iloc[0] if not grand_total.empty else None

    records: list[dict[str, Any]] = []
    for row_value in row_values:
        row_key = _label_key(row_value)
        record: dict[str, Any] = {row_bin: _label_display(row_value)}
        for column_value in column_values:
            column_key = _label_key(column_value)
            record[_label_display(column_value)] = _format_cell(
                detail_records.get((row_key, column_key)), metric
            )
        record["Total"] = _format_cell(row_total_records.get((row_key,)), metric)
        records.append(record)

    total_record: dict[str, Any] = {row_bin: "Total"}
    for column_value in column_values:
        column_key = _label_key(column_value)
        total_record[_label_display(column_value)] = _format_cell(
            column_total_records.get((column_key,)), metric
        )
    total_record["Total"] = _format_cell(grand_record, metric)
    records.append(total_record)
    return pd.DataFrame(records)


def _user_profile_cfg_with_numeric_metrics(
    cfg: dict[str, Any],
    metrics: list[dict[str, Any]],
) -> dict[str, Any]:
    user_profile_metrics = dict(cfg.get("user_profile_metrics", {}))
    user_profile_metrics["numeric_cross_metrics"] = metrics
    out = dict(cfg)
    out["user_profile_metrics"] = user_profile_metrics
    return out


def _valid_user_profile_numeric_metrics(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    logger,
) -> list[dict[str, Any]]:
    valid: list[dict[str, Any]] = []
    for metric_cfg in cfg.get("user_profile_metrics", {}).get("numeric_cross_metrics", []):
        source_field = metric_cfg.get("source_field")
        if not source_field:
            if logger:
                logger.warning(f"User profile numeric metric missing source_field, skipped: {metric_cfg}")
            continue
        if source_field not in df.columns:
            if logger:
                logger.warning(f"User profile numeric field missing, skipped: {source_field}")
            continue
        agg = str(metric_cfg.get("agg", "mean")).lower()
        if agg != "mean":
            if logger:
                logger.warning(f"Unsupported user profile numeric agg skipped: field={source_field}, agg={agg}")
            continue
        valid.append(metric_cfg)
    return valid


def _valid_user_profile_category_metrics(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    logger,
) -> list[dict[str, Any]]:
    valid: list[dict[str, Any]] = []
    for metric_cfg in cfg.get("user_profile_metrics", {}).get("category_distribution_metrics", []):
        source_field = metric_cfg.get("source_field")
        if not source_field:
            if logger:
                logger.warning(f"User profile category metric missing source_field, skipped: {metric_cfg}")
            continue
        if source_field not in df.columns:
            if logger:
                logger.warning(f"User profile category field missing, skipped: {source_field}")
            continue
        valid.append(metric_cfg)
    return valid


def _build_user_profile_numeric_pivot(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    metric: str,
    numeric_metrics: list[dict[str, Any]],
) -> pd.DataFrame:
    numeric_cfg = _user_profile_cfg_with_numeric_metrics(cfg, numeric_metrics)
    detail = calculate_user_profile_numeric_metrics(df, [row_bin, column_bin], numeric_cfg)
    row_totals = calculate_user_profile_numeric_metrics(df, [row_bin], numeric_cfg)
    column_totals = calculate_user_profile_numeric_metrics(df, [column_bin], numeric_cfg)
    grand_total = calculate_user_profile_numeric_metrics(df, [], numeric_cfg)

    detail_records = _records_by_key(detail, [row_bin, column_bin])
    row_total_records = _records_by_key(row_totals, [row_bin])
    column_total_records = _records_by_key(column_totals, [column_bin])
    grand_record = grand_total.iloc[0] if not grand_total.empty else None

    records: list[dict[str, Any]] = []
    for row_value in USER_PROFILE_BIN_ORDER:
        row_key = _label_key(row_value)
        record: dict[str, Any] = {row_bin: _label_display(row_value)}
        for column_value in USER_PROFILE_BIN_ORDER:
            column_key = _label_key(column_value)
            cell_record = detail_records.get((row_key, column_key))
            record[_label_display(column_value)] = (
                "" if cell_record is None else _format_user_profile_value(cell_record.get(metric), metric)
            )
        row_total_record = row_total_records.get((row_key,))
        record["Total"] = (
            "" if row_total_record is None else _format_user_profile_value(row_total_record.get(metric), metric)
        )
        records.append(record)

    total_record: dict[str, Any] = {row_bin: "Total"}
    for column_value in USER_PROFILE_BIN_ORDER:
        column_key = _label_key(column_value)
        column_total_record = column_total_records.get((column_key,))
        total_record[_label_display(column_value)] = (
            "" if column_total_record is None else _format_user_profile_value(column_total_record.get(metric), metric)
        )
    total_record["Total"] = "" if grand_record is None else _format_user_profile_value(grand_record.get(metric), metric)
    records.append(total_record)
    return pd.DataFrame(records)


def _empty_user_profile_numeric_row(numeric_metrics: list[dict[str, Any]]) -> dict[str, Any]:
    row: dict[str, Any] = {"sample_cnt": 0}
    for metric_cfg in numeric_metrics:
        metric = metric_cfg.get("output_field", f"avg_{metric_cfg.get('source_field', '')}")
        row[metric] = None
    return row


def _complete_user_profile_numeric_metrics(
    metrics: pd.DataFrame,
    numeric_metrics: list[dict[str, Any]],
    group_cols: list[str],
) -> pd.DataFrame:
    if not group_cols:
        return metrics

    expected_values = {col: USER_PROFILE_BIN_ORDER for col in group_cols}
    existing = _records_by_key(metrics, group_cols) if not metrics.empty else {}
    records: list[dict[str, Any]] = []

    if len(group_cols) == 1:
        col = group_cols[0]
        for value in expected_values[col]:
            key = (_label_key(value),)
            if key in existing:
                records.append(existing[key].to_dict())
            else:
                row = {col: value}
                row.update(_empty_user_profile_numeric_row(numeric_metrics))
                records.append(row)
    elif len(group_cols) == 2:
        row_col, column_col = group_cols
        for row_value in expected_values[row_col]:
            for column_value in expected_values[column_col]:
                key = (_label_key(row_value), _label_key(column_value))
                if key in existing:
                    records.append(existing[key].to_dict())
                else:
                    row = {row_col: row_value, column_col: column_value}
                    row.update(_empty_user_profile_numeric_row(numeric_metrics))
                    records.append(row)
    else:
        return metrics

    columns = group_cols + ["sample_cnt"] + [
        metric_cfg.get("output_field", f"avg_{metric_cfg.get('source_field', '')}")
        for metric_cfg in numeric_metrics
    ]
    return pd.DataFrame(records).reindex(columns=columns)


def _category_metric_groups(metrics: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for metric_cfg in metrics:
        grouped.setdefault(metric_cfg.get("sheet", "category_distribution"), []).append(metric_cfg)

    ordered: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in USER_PROFILE_DISTRIBUTION_SHEET_ORDER:
        if sheet_name in grouped:
            ordered[sheet_name] = grouped.pop(sheet_name)
    ordered.update(grouped)
    return ordered


def _build_category_distribution_pivot(
    distribution: pd.DataFrame,
    field: str,
    value_col: str,
) -> pd.DataFrame:
    field_data = distribution[distribution["profile_field"].eq(field)].copy()
    category_order = (
        field_data[
            field_data["primary_model_score_bin"].astype(str).eq("Total")
            & ~field_data["category_value"].astype(str).eq("Total")
        ]["category_value"]
        .astype(str)
        .tolist()
    )
    columns = category_order + ["Total"]
    records: list[dict[str, Any]] = []

    for row_value in USER_PROFILE_BIN_ORDER + ["Total"]:
        row_data = field_data[field_data["primary_model_score_bin"].astype(str).eq(str(row_value))]
        row_record: dict[str, Any] = {"primary_model_score_bin": _label_display(row_value)}
        for category_value in columns:
            match = row_data[row_data["category_value"].astype(str).eq(str(category_value))]
            if match.empty:
                row_record[category_value] = ""
                continue
            value = match.iloc[0][value_col]
            metric_name = "category_pct" if value_col.endswith("_pct") else "category_cnt"
            row_record[category_value] = _format_user_profile_value(value, metric_name)
        records.append(row_record)

    return pd.DataFrame(records, columns=["primary_model_score_bin"] + columns)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int) -> tuple[int, int]:
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            ws.cell(
                row=start_row + row_offset,
                column=1 + col_offset,
                value=None if _is_missing(value) else value,
            )
    return len(rows), len(df.columns)


def _style_table(ws, start_row: int, row_count: int, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    end_row = start_row + row_count - 1

    for col in range(1, col_count + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in range(start_row + 1, end_row + 1):
        is_total_row = ws.cell(row=row, column=1).value == "Total"
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if is_total_row or ws.cell(row=start_row, column=col).value == "Total":
                cell.fill = total_fill


def _auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for item in cell:
                if item.value is not None:
                    max_len = max(max_len, len(str(item.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 32)


def write_cross_model_workbook(
    output_path: Path,
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    logger,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "metric_guide"

    guide_df = pd.DataFrame(
        [
            {
                "category": category,
                "metric": metric,
                "cell_format": "percent" if metric in RATE_DENOMINATORS else "value",
                "denominator_metric": RATE_DENOMINATORS.get(metric, ""),
            }
            for category, metrics in METRIC_GROUPS.items()
            for metric in metrics
        ]
    )
    row_count, col_count = _write_dataframe(ws, guide_df, 1)
    _style_table(ws, 1, row_count, col_count)
    ws.freeze_panes = "A2"
    _auto_width(ws)

    for category, metrics in METRIC_GROUPS.items():
        ws = wb.create_sheet(category)
        current_row = 1
        for metric in metrics:
            ws.cell(row=current_row, column=1, value=metric).font = Font(bold=True, size=12)
            current_row += 1
            pivot = _build_metric_pivot(df, cfg, row_bin, column_bin, metric)
            row_count, col_count = _write_dataframe(ws, pivot, current_row)
            _style_table(ws, current_row, row_count, col_count)
            current_row += row_count + 2
        _auto_width(ws)

    raw_outputs = {
        "raw_cross_metrics": calculate_group_metrics(df, [row_bin, column_bin], cfg),
        "raw_row_totals": calculate_group_metrics(df, [row_bin], cfg),
        "raw_column_totals": calculate_group_metrics(df, [column_bin], cfg),
        "raw_grand_total": calculate_group_metrics(df, [], cfg),
    }
    for sheet_name, data in raw_outputs.items():
        ws = wb.create_sheet(sheet_name)
        row_count, col_count = _write_dataframe(ws, data, 1)
        _style_table(ws, 1, row_count, col_count)
        ws.freeze_panes = "C2" if sheet_name == "raw_cross_metrics" else "A2"
        _auto_width(ws)

    wb.save(output_path)
    logger.info(f"Wrote workbook: {output_path}")


def export_cross_model_bin_user_profile_excel(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    output_path: Path,
    row_bin: str = "primary_model_score_bin",
    column_bin: str = "comparison_model_score_bin",
    logger=None,
) -> None:
    numeric_metrics = _valid_user_profile_numeric_metrics(df, cfg, logger)
    category_metrics = _valid_user_profile_category_metrics(df, cfg, logger)
    numeric_cfg = _user_profile_cfg_with_numeric_metrics(cfg, numeric_metrics)

    wb = Workbook()
    ws = wb.active
    ws.title = "metric_guide"

    guide_df = _user_profile_metric_guide(numeric_metrics, category_metrics)
    row_count, col_count = _write_dataframe(ws, guide_df, 1)
    _style_table(ws, 1, row_count, col_count)
    _apply_metric_number_format(ws, 1, row_count, list(guide_df.columns))
    ws.freeze_panes = "A2"
    _auto_width(ws)

    numeric_groups = _user_profile_numeric_metric_groups(numeric_metrics)
    category_groups = _category_metric_groups(category_metrics)

    def write_numeric_sheet(category: str) -> None:
        if category not in numeric_groups:
            return
        ws = wb.create_sheet(category)
        ws.freeze_panes = "A2"
        current_row = 1
        for metric in numeric_groups[category]:
            ws.cell(row=current_row, column=1, value=metric).font = Font(bold=True, size=12)
            current_row += 1
            pivot = _build_user_profile_numeric_pivot(df, cfg, row_bin, column_bin, metric, numeric_metrics)
            row_count, col_count = _write_dataframe(ws, pivot, current_row)
            _style_table(ws, current_row, row_count, col_count)
            current_row += row_count + 2
        _auto_width(ws)

    distribution_frames = [
        calculate_user_profile_category_distribution(df, row_bin, metric_cfg, USER_PROFILE_BIN_ORDER)
        for metric_cfg in category_metrics
    ]
    distribution = (
        pd.concat(distribution_frames, ignore_index=True)
        if distribution_frames
        else pd.DataFrame(
            columns=[
                "profile_field",
                "primary_model_score_bin",
                "category_value",
                "sample_cnt",
                "category_cnt",
                "category_pct",
            ]
        )
    )

    def write_distribution_sheet(sheet_name: str) -> None:
        if sheet_name not in category_groups:
            return
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A2"
        current_row = 1
        for metric_cfg in category_groups[sheet_name]:
            field = metric_cfg["source_field"]
            for suffix, value_col in [("cnt", "category_cnt"), ("pct", "category_pct")]:
                metric_name = f"{field}_{suffix}"
                ws.cell(row=current_row, column=1, value=metric_name).font = Font(bold=True, size=12)
                current_row += 1
                pivot = _build_category_distribution_pivot(distribution, field, value_col)
                row_count, col_count = _write_dataframe(ws, pivot, current_row)
                _style_table(ws, current_row, row_count, col_count)
                current_row += row_count + 2
        _auto_width(ws)

    write_numeric_sheet("basic_profile")
    write_distribution_sheet("geo_profile_distribution")
    write_distribution_sheet("family_profile_distribution")
    write_distribution_sheet("attributed_category_distribution")
    write_numeric_sheet("occupation_profile")
    write_distribution_sheet("occupation_profile_distribution")

    raw_numeric = _complete_user_profile_numeric_metrics(
        calculate_user_profile_numeric_metrics(df, [row_bin, column_bin], numeric_cfg),
        numeric_metrics,
        [row_bin, column_bin],
    )
    raw_outputs = {
        "raw_user_profile_numeric_cross_metrics": raw_numeric,
        "raw_user_profile_category_distribution": distribution,
    }
    for sheet_name, data in raw_outputs.items():
        ws = wb.create_sheet(sheet_name)
        row_count, col_count = _write_dataframe(ws, data, 1)
        _style_table(ws, 1, row_count, col_count)
        _apply_metric_number_format(ws, 1, row_count, list(data.columns))
        ws.freeze_panes = "A2"
        _auto_width(ws)

    wb.save(output_path)
    if logger:
        logger.info(f"Wrote user profile workbook: {output_path}")
