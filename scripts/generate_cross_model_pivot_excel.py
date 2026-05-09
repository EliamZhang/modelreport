from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.aggregation import enrich_base_sample
from src.binning import apply_score_binning
from src.config_loader import get_output_dir, load_config
from src.data_loader import load_input_tables
from src.label_builder import apply_deal_amount_filter, build_conversion_labels
from src.logger import setup_logger
from src.metric_calculator import calculate_group_metrics
from src.monthly_analysis import add_sample_month


METRIC_GROUPS: dict[str, list[str]] = {
    "risk_perf": [
        "duedate_3m_30_valid_cnt",
        "duedate_3m_30_bad_cnt",
        "duedate_3m_30_good_cnt",
        "duedate_3m_30_bad_rate",
        "duedate_1m_5_valid_cnt",
        "duedate_1m_5_bad_cnt",
        "duedate_1m_5_good_cnt",
        "duedate_1m_5_bad_rate",
    ],
    "amount_risk": [
        "duedate_3m_30_amount_overdue_amount",
        "duedate_3m_30_amount_principal_amount",
        "duedate_3m_30_amount_overdue_rate",
        "duedate_1m_5_amount_overdue_amount",
        "duedate_1m_5_amount_principal_amount",
        "duedate_1m_5_amount_overdue_rate",
    ],
    "profile": [
        "avg_PTI",
        "avg_requested_loan_amount",
        "avg_total_amount",
        "avg_gross_surplus",
        "avg_net_surplus",
        "avg_total_income",
    ],
    "conversion": [
        "apply_cnt",
        "completed_application_cnt",
        "approved_application_cnt",
        "auto_approved_application_cnt",
        "manual_approved_application_cnt",
        "deal_sample_cnt",
        "completion_rate",
        "approval_rate",
        "auto_approval_rate",
        "manual_approval_rate",
        "auto_approval_share",
        "manual_approval_share",
        "deal_rate",
    ],
}

RATE_DENOMINATORS = {
    "duedate_3m_30_bad_rate": "duedate_3m_30_valid_cnt",
    "duedate_1m_5_bad_rate": "duedate_1m_5_valid_cnt",
    "duedate_3m_30_amount_overdue_rate": "duedate_3m_30_amount_principal_amount",
    "duedate_1m_5_amount_overdue_rate": "duedate_1m_5_amount_principal_amount",
    "completion_rate": "apply_cnt",
    "approval_rate": "completed_application_cnt",
    "auto_approval_rate": "completed_application_cnt",
    "manual_approval_rate": "completed_application_cnt",
    "auto_approval_share": "approved_application_cnt",
    "manual_approval_share": "approved_application_cnt",
    "deal_rate": "approved_application_cnt",
}

COUNT_METRICS = {
    "sample_cnt",
    "duedate_3m_30_valid_cnt",
    "duedate_3m_30_bad_cnt",
    "duedate_3m_30_good_cnt",
    "duedate_1m_5_valid_cnt",
    "duedate_1m_5_bad_cnt",
    "duedate_1m_5_good_cnt",
    "apply_cnt",
    "completed_application_cnt",
    "approved_application_cnt",
    "auto_approved_application_cnt",
    "manual_approved_application_cnt",
    "deal_sample_cnt",
}

AMOUNT_METRICS = {
    "duedate_3m_30_amount_overdue_amount",
    "duedate_3m_30_amount_principal_amount",
    "duedate_1m_5_amount_overdue_amount",
    "duedate_1m_5_amount_principal_amount",
}

FOUR_DECIMAL_METRICS = {
    "avg_PTI",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an Excel workbook of primary x comparison model-bin pivot tables."
    )
    parser.add_argument(
        "--config",
        default=str(PROJECT_ROOT / "config" / "analysis_config_input_sample.py"),
        help="Path to the existing analysis config.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Workbook path. Relative paths are written under the configured output_dir.",
    )
    parser.add_argument(
        "--row-bin",
        default="primary_model_score_bin",
        help="Row bin field for the pivot tables.",
    )
    parser.add_argument(
        "--column-bin",
        default="comparison_model_score_bin",
        help="Column bin field for the pivot tables.",
    )
    return parser.parse_args()


def _resolve_output_path(output_arg: str | None, cfg: dict[str, Any]) -> Path:
    out_dir = get_output_dir(cfg)
    if not output_arg:
        return out_dir / "cross_model_bin_pivot_rate_only.xlsx"
    output_path = Path(output_arg).expanduser()
    if output_path.is_absolute():
        return output_path
    return out_dir / output_path


def _prepare_output_dir(cfg: dict[str, Any]) -> Path:
    out_dir = get_output_dir(cfg)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not cfg.get("runtime", {}).get("overwrite_output", True):
        return out_dir

    project_root = Path(cfg.get("project", {}).get("root_dir", ".")).expanduser().resolve()
    try:
        out_dir.relative_to(project_root)
    except ValueError as exc:
        raise ValueError(
            f"Refuse to clear output_dir outside project root when overwrite_output=True: {out_dir}"
        ) from exc

    for child in out_dir.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()

    return out_dir


def _build_enriched_sample(cfg: dict[str, Any], logger) -> pd.DataFrame:
    tables = load_input_tables(cfg, logger)
    enriched, _ = enrich_base_sample(tables, cfg, logger)
    enriched = apply_score_binning(enriched, cfg, logger)
    enriched = add_sample_month(enriched, cfg, logger)
    enriched = build_conversion_labels(enriched, cfg, logger)
    enriched = apply_deal_amount_filter(enriched, cfg, logger)
    return enriched


def _is_missing(value: Any) -> bool:
    return value is None or pd.isna(value)


def _label_key(value: Any) -> str:
    if _is_missing(value):
        return "__NA__"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _label_display(value: Any) -> str:
    if _is_missing(value):
        return "NA"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sort_values(values: list[Any]) -> list[Any]:
    def sort_key(value: Any) -> tuple[int, float, str]:
        if _is_missing(value):
            return (2, 0.0, "")
        try:
            return (0, float(value), str(value))
        except (TypeError, ValueError):
            return (1, 0.0, str(value))

    unique: dict[str, Any] = {}
    for value in values:
        unique.setdefault(_label_key(value), value)
    return sorted(unique.values(), key=sort_key)


def _records_by_key(metrics: pd.DataFrame, group_cols: list[str]) -> dict[tuple[str, ...], pd.Series]:
    records: dict[tuple[str, ...], pd.Series] = {}
    for _, row in metrics.iterrows():
        key = tuple(_label_key(row[col]) for col in group_cols)
        records[key] = row
    return records


def _format_number(value: Any, metric: str) -> str:
    if _is_missing(value):
        return ""
    value = float(value)
    if metric in COUNT_METRICS:
        return f"{int(round(value)):,}"
    if metric in AMOUNT_METRICS:
        return f"{value:,.2f}"
    if metric in FOUR_DECIMAL_METRICS:
        return f"{value:.4f}"
    if metric.startswith("avg_"):
        return f"{value:,.2f}"
    return f"{value:,.4f}"


def _format_cell(record: pd.Series | None, metric: str) -> str:
    if record is None or metric not in record.index:
        return ""

    value = record[metric]
    if metric in RATE_DENOMINATORS:
        if _is_missing(value):
            return ""
        return f"{float(value):.2%}"

    return _format_number(value, metric)


def build_metric_pivot(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    metric: str,
) -> pd.DataFrame:
    detail = calculate_group_metrics(df, [row_bin, column_bin], cfg)
    row_totals = calculate_group_metrics(df, [row_bin], cfg)
    column_totals = calculate_group_metrics(df, [column_bin], cfg)
    grand_total = calculate_group_metrics(df, [], cfg)

    row_values = _sort_values(list(detail[row_bin]) + list(row_totals[row_bin]))
    column_values = _sort_values(list(detail[column_bin]) + list(column_totals[column_bin]))

    detail_records = _records_by_key(detail, [row_bin, column_bin])
    row_total_records = _records_by_key(row_totals, [row_bin])
    column_total_records = _records_by_key(column_totals, [column_bin])
    grand_record = grand_total.iloc[0] if not grand_total.empty else None

    records: list[dict[str, Any]] = []
    for row_value in row_values:
        row_key = _label_key(row_value)
        record: dict[str, Any] = {row_bin: _label_display(row_value)}
        for column_value in column_values:
            column_key = _label_key(column_value)
            record[_label_display(column_value)] = _format_cell(
                detail_records.get((row_key, column_key)),
                metric,
            )
        record["Total"] = _format_cell(row_total_records.get((row_key,)), metric)
        records.append(record)

    total_record: dict[str, Any] = {row_bin: "Total"}
    for column_value in column_values:
        column_key = _label_key(column_value)
        total_record[_label_display(column_value)] = _format_cell(
            column_total_records.get((column_key,)),
            metric,
        )
    total_record["Total"] = _format_cell(grand_record, metric)
    records.append(total_record)

    return pd.DataFrame(records)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> tuple[int, int]:
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            if _is_missing(value):
                value = None
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
    return len(rows), len(df.columns)


def _style_table(ws, start_row: int, row_count: int, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    end_row = start_row + row_count - 1
    end_col = col_count

    for col in range(1, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in range(start_row + 1, end_row + 1):
        is_total_row = ws.cell(row=row, column=1).value == "Total"
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if is_total_row or ws.cell(row=start_row, column=col).value == "Total":
                cell.fill = total_fill


def _auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for item in cell:
                if item.value is not None:
                    max_len = max(max_len, len(str(item.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 32)


def write_workbook(
    output_path: Path,
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    logger,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    raw_cross = calculate_group_metrics(df, [row_bin, column_bin], cfg)
    row_totals = calculate_group_metrics(df, [row_bin], cfg)
    column_totals = calculate_group_metrics(df, [column_bin], cfg)
    grand_total = calculate_group_metrics(df, [], cfg)

    guide_records = []
    for category, metrics in METRIC_GROUPS.items():
        for metric in metrics:
            guide_records.append(
                {
                    "category": category,
                    "metric": metric,
                    "cell_format": "percent" if metric in RATE_DENOMINATORS else "value",
                    "denominator_metric": RATE_DENOMINATORS.get(metric, ""),
                }
            )
    guide_df = pd.DataFrame(guide_records)

    wb = Workbook()
    ws = wb.active
    ws.title = "metric_guide"
    row_count, col_count = _write_dataframe(ws, guide_df, start_row=1)
    _style_table(ws, 1, row_count, col_count)
    ws.freeze_panes = "A2"
    _auto_width(ws)

    for category, metrics in METRIC_GROUPS.items():
        ws = wb.create_sheet(category)
        current_row = 1
        for metric in metrics:
            ws.cell(row=current_row, column=1, value=metric)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            pivot = build_metric_pivot(df, cfg, row_bin, column_bin, metric)
            row_count, col_count = _write_dataframe(ws, pivot, start_row=current_row)
            _style_table(ws, current_row, row_count, col_count)
            current_row += row_count + 2
        _auto_width(ws)

    for sheet_name, data in {
        "raw_cross_metrics": raw_cross,
        "raw_row_totals": row_totals,
        "raw_column_totals": column_totals,
        "raw_grand_total": grand_total,
    }.items():
        ws = wb.create_sheet(sheet_name)
        row_count, col_count = _write_dataframe(ws, data, start_row=1)
        if row_count and col_count:
            _style_table(ws, 1, row_count, col_count)
        ws.freeze_panes = "C2" if sheet_name == "raw_cross_metrics" else "A2"
        _auto_width(ws)

    wb.save(output_path)
    logger.info(f"Wrote cross model pivot workbook: {output_path}")


def main() -> None:
    args = parse_args()
    cfg = load_config(args.config)
    _prepare_output_dir(cfg)
    output_path = _resolve_output_path(args.output, cfg)
    logger = setup_logger(
        output_path.parent,
        log_file=None,
        level=cfg.get("runtime", {}).get("log_level", "INFO"),
    )

    logger.info("Cross model pivot workbook generation started")
    logger.info(f"Config path: {Path(args.config).resolve()}")
    logger.info(f"Output path: {output_path}")
    enriched = _build_enriched_sample(cfg, logger)

    missing_bins = [col for col in [args.row_bin, args.column_bin] if col not in enriched.columns]
    if missing_bins:
        raise ValueError(f"Pivot bin columns missing after enrichment: {missing_bins}")

    write_workbook(output_path, enriched, cfg, args.row_bin, args.column_bin, logger)
    logger.info("Cross model pivot workbook generation finished")


if __name__ == "__main__":
    main()
