from __future__ import annotations

import logging
import shutil
import sys
from copy import deepcopy
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


METRIC_GROUPS: dict[str, list[str]] = {
    "risk_perf": [
        "duedate_3m_30_valid_cnt",
        "duedate_3m_30_bad_cnt",
        "duedate_3m_30_good_cnt",
        "duedate_3m_30_bad_rate",
        "duedate_1m_5_valid_cnt",
        "duedate_1m_5_bad_cnt",
        "duedate_1m_5_good_cnt",
        "duedate_1m_5_bad_rate",
    ],
    "amount_risk": [
        "duedate_3m_30_amount_overdue_amount",
        "duedate_3m_30_amount_principal_amount",
        "duedate_3m_30_amount_overdue_rate",
        "duedate_1m_5_amount_overdue_amount",
        "duedate_1m_5_amount_principal_amount",
        "duedate_1m_5_amount_overdue_rate",
    ],
    "profile": [
        "avg_PTI",
        "avg_requested_loan_amount",
        "avg_total_amount",
        "avg_gross_surplus",
        "avg_net_surplus",
        "avg_total_income",
    ],
    "conversion": [
        "apply_cnt",
        "completed_application_cnt",
        "approved_application_cnt",
        "auto_approved_application_cnt",
        "manual_approved_application_cnt",
        "deal_sample_cnt",
        "completion_rate",
        "approval_rate",
        "auto_approval_rate",
        "manual_approval_rate",
        "auto_approval_share",
        "manual_approval_share",
        "deal_rate",
    ],
}

RATE_DENOMINATORS = {
    "duedate_3m_30_bad_rate": "duedate_3m_30_valid_cnt",
    "duedate_1m_5_bad_rate": "duedate_1m_5_valid_cnt",
    "duedate_3m_30_amount_overdue_rate": "duedate_3m_30_amount_principal_amount",
    "duedate_1m_5_amount_overdue_rate": "duedate_1m_5_amount_principal_amount",
    "completion_rate": "apply_cnt",
    "approval_rate": "completed_application_cnt",
    "auto_approval_rate": "completed_application_cnt",
    "manual_approval_rate": "completed_application_cnt",
    "auto_approval_share": "approved_application_cnt",
    "manual_approval_share": "approved_application_cnt",
    "deal_rate": "approved_application_cnt",
}

COUNT_METRICS = {
    "sample_cnt",
    "duedate_3m_30_valid_cnt",
    "duedate_3m_30_bad_cnt",
    "duedate_3m_30_good_cnt",
    "duedate_1m_5_valid_cnt",
    "duedate_1m_5_bad_cnt",
    "duedate_1m_5_good_cnt",
    "apply_cnt",
    "completed_application_cnt",
    "approved_application_cnt",
    "auto_approved_application_cnt",
    "manual_approved_application_cnt",
    "deal_sample_cnt",
}

AMOUNT_METRICS = {
    "duedate_3m_30_amount_overdue_amount",
    "duedate_3m_30_amount_principal_amount",
    "duedate_1m_5_amount_overdue_amount",
    "duedate_1m_5_amount_principal_amount",
}

FOUR_DECIMAL_METRICS = {"avg_PTI"}
USER_PROFILE_CATEGORY_ORDER = ["basic_profile", "occupation_profile"]
USER_PROFILE_DISTRIBUTION_SHEET_ORDER = [
    "geo_profile_distribution",
    "family_profile_distribution",
    "attributed_category_distribution",
    "occupation_profile_distribution",
]
VARIABLE_NAME_COLUMNS = ["field", "fields", "variable", "variable_name", "metric", "feature", "feature_name"]
KEY_COLUMNS = {"application_id", "user_id", "sample_datetime"}
MAX_EXCEL_ROWS = 1_048_576
EXCEL_FONT_NAME = "Microsoft YaHei"


def setup_logger(level: str = "INFO") -> logging.Logger:
    logger = logging.getLogger("model_analysis")
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))
    logger.propagate = False
    logger.handlers.clear()

    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(getattr(logging, level.upper(), logging.INFO))
    handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", "%Y-%m-%d %H:%M:%S"))
    logger.addHandler(handler)
    return logger


def resolve_config_paths(config: dict[str, Any]) -> dict[str, Any]:
    cfg = deepcopy(config)
    project = cfg["project"]
    root_dir = Path(project.get("root_dir", ".")).expanduser()
    if not root_dir.is_absolute():
        root_dir = Path.cwd() / root_dir
    project["root_dir"] = str(root_dir.resolve())

    output_dir = Path(project["output_dir"]).expanduser()
    if not output_dir.is_absolute():
        output_dir = Path(project["root_dir"]) / output_dir
    project["output_dir"] = str(output_dir.resolve())

    for table_cfg in cfg["input_tables"].values():
        path = Path(table_cfg["path"]).expanduser()
        if not path.is_absolute():
            path = Path(project["root_dir"]) / path
        table_cfg["path"] = str(path.resolve())

    feature_cfg = cfg.setdefault("feature_profile", {})
    library_path = Path(feature_cfg.get("variable_library_path", "./INPUT/model_variable_library.csv")).expanduser()
    if not library_path.is_absolute():
        library_path = Path(project["root_dir"]) / library_path
    feature_cfg["variable_library_path"] = str(library_path.resolve())
    return cfg


def prepare_output_dir(cfg: dict[str, Any]) -> Path:
    out_dir = Path(cfg["project"]["output_dir"]).resolve()
    project_root = Path(cfg["project"]["root_dir"]).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        out_dir.relative_to(project_root)
    except ValueError as exc:
        raise ValueError(f"Refuse to clear output_dir outside project root: {out_dir}") from exc

    for child in out_dir.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()
    return out_dir


def _normalize_key_series(series: pd.Series) -> pd.Series:
    normalized = series.astype("string").str.strip()
    normalized = normalized.str.replace(r"\.0+$", "", regex=True)
    return normalized.mask(normalized.isin(["", "nan", "None", "<NA>"]), pd.NA)


def _key_columns(cfg: dict[str, Any]) -> set[str]:
    keys = cfg["keys"]
    columns = {keys["primary_key"], keys["user_key"]}
    columns.update(join_cfg["join_key"] for join_cfg in cfg["joins"].values())
    return columns


def _score_binning_schemes(cfg: dict[str, Any]) -> list[dict[str, Any]]:
    binning_cfg = cfg.get("score_binning", {})
    schemes = binning_cfg.get("schemes")
    if schemes:
        return list(schemes)
    return [
        {
            "name": binning_cfg.get("default_scheme", "auto"),
            "title": binning_cfg.get("title", binning_cfg.get("default_scheme", "auto")),
            "binning_mode": binning_cfg.get("binning_mode", "equal_frequency"),
            "scores": binning_cfg.get("scores", []),
        }
    ]


def _score_binning_scheme_by_name(cfg: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {str(scheme.get("name")): scheme for scheme in _score_binning_schemes(cfg)}


def _required_columns(cfg: dict[str, Any]) -> dict[str, set[str] | None]:
    required: dict[str, set[str] | None] = {cfg["analysis"]["base_table"]: None}
    for table_name in cfg["input_tables"]:
        required.setdefault(table_name, set(_key_columns(cfg)))

    for scheme_cfg in _score_binning_schemes(cfg):
        for score_cfg in scheme_cfg.get("scores", []):
            table_name = score_cfg["source_table"]
            if required.get(table_name) is not None:
                required[table_name].add(score_cfg["score_field"])

    for join_cfg in cfg["joins"].values():
        table_name = join_cfg["source_table"]
        if required.get(table_name) is not None:
            required[table_name].add(join_cfg["join_key"])
            required[table_name].update(join_cfg["fields"])
    return required


def load_input_tables(cfg: dict[str, Any], logger: logging.Logger) -> dict[str, pd.DataFrame]:
    key_columns = _key_columns(cfg)
    required = _required_columns(cfg)
    tables: dict[str, pd.DataFrame] = {}

    for table_name, table_cfg in cfg["input_tables"].items():
        columns = required.get(table_name)
        path = Path(table_cfg["path"])
        logger.info(f"Loading table={table_name}, path={path}, columns={'ALL' if columns is None else len(columns)}")
        usecols = None if columns is None else lambda col_name: str(col_name) in columns
        df = pd.read_csv(path, low_memory=False, usecols=usecols)
        for col in key_columns:
            if col in df.columns:
                df[col] = _normalize_key_series(df[col])
        tables[table_name] = df
        logger.info(f"Loaded table={table_name}, rows={len(df):,}, cols={len(df.columns):,}")
    return tables


def _merge_with_overwrite(base: pd.DataFrame, right: pd.DataFrame, key: str, source_table: str, logger) -> pd.DataFrame:
    overlap_cols = [col for col in right.columns if col != key and col in base.columns]
    if not overlap_cols:
        return base.merge(right, on=key, how="left")

    logger.warning(f"Join overwrite columns detected: source_table={source_table}, columns={overlap_cols}")
    renamed = {col: f"__overwrite__{col}" for col in overlap_cols}
    marker = "__right_matched__"
    right_to_merge = right.rename(columns=renamed).copy()
    right_to_merge[marker] = True
    merged = base.merge(right_to_merge, on=key, how="left")
    matched = merged[marker].eq(True)

    for col in overlap_cols:
        incoming_col = renamed[col]
        merged[col] = merged[incoming_col].where(matched, merged[col])
        merged = merged.drop(columns=[incoming_col])
    return merged.drop(columns=[marker])


def enrich_base_sample(tables: dict[str, pd.DataFrame], cfg: dict[str, Any], logger) -> pd.DataFrame:
    key = cfg["keys"]["primary_key"]
    base_table = cfg["analysis"]["base_table"]
    base = tables[base_table].copy()
    if key not in base.columns:
        raise ValueError(f"Base table missing key column: {key}")

    logger.info(f"Base sample rows={len(base):,}, cols={len(base.columns):,}")
    for join_cfg in cfg["joins"].values():
        source_table = join_cfg["source_table"]
        join_key = join_cfg["join_key"]
        if join_key != key:
            raise ValueError(f"Join key must match primary key. primary_key={key}, join_key={join_key}")

        right = tables[source_table][[join_key] + join_cfg["fields"]].copy()
        duplicate_rows = int(right.duplicated(subset=[join_key], keep=False).sum())
        if duplicate_rows:
            logger.warning(f"Join table has duplicate keys; keeping first: table={source_table}, rows={duplicate_rows:,}")
            right = right.drop_duplicates(subset=[join_key], keep="first")

        before_rows = len(base)
        base = _merge_with_overwrite(base, right, key, source_table, logger)
        if len(base) != before_rows:
            raise ValueError(f"Join changed row count: table={source_table}, before={before_rows}, after={len(base)}")
        logger.info(f"Joined table={source_table}, rows={len(base):,}, cols={len(base.columns):,}")
    return base


def _cast_bin_label(series: pd.Series, label_type: str | None) -> pd.Series:
    if label_type == "int":
        return pd.to_numeric(series, errors="coerce").astype("Int64")
    if label_type == "float":
        return pd.to_numeric(series, errors="coerce")
    return series


def _coerce_numeric_values(values: list[Any] | None) -> list[float]:
    if not values:
        return []
    return pd.to_numeric(pd.Series(values), errors="coerce").dropna().tolist()


def _quantile_bin_labels(score_cfg: dict[str, Any], label_type: str | None, bin_count: int) -> list[Any]:
    configured_labels = score_cfg.get("bin_labels")
    if configured_labels:
        labels = list(configured_labels)
    elif label_type == "float":
        labels = [float(i) for i in range(1, bin_count + 1)]
    else:
        labels = list(range(1, bin_count + 1))
    if len(labels) != bin_count:
        raise ValueError(f"bin_labels length must equal bin_count={bin_count}, got {len(labels)}")
    return labels


def _special_label_map(score_cfg: dict[str, Any]) -> dict[float, Any]:
    return {float(key): value for key, value in score_cfg.get("special_label_map", {}).items()}


def _normalize_bin_label(value: Any) -> str:
    if value is None or pd.isna(value):
        return "__NA__"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _resolve_group_indexes(group_cfg: Any, bins: list[dict[str, Any]]) -> tuple[Any, list[int]]:
    group_label = None
    source_values = group_cfg
    source_mode = "auto"

    if isinstance(group_cfg, dict):
        group_label = group_cfg.get("label")
        if "source_bin_indexes" in group_cfg:
            source_values = group_cfg.get("source_bin_indexes")
            source_mode = "index"
        elif "source_bin_labels" in group_cfg:
            source_values = group_cfg.get("source_bin_labels")
            source_mode = "label"
        elif "bins" in group_cfg:
            source_values = group_cfg.get("bins")

    if not isinstance(source_values, (list, tuple)) or not source_values:
        raise ValueError(f"Each bin group must contain at least one source bin, got {group_cfg}")

    label_to_index = {_normalize_bin_label(bin_cfg.get("label")): index for index, bin_cfg in enumerate(bins)}
    values = list(source_values)
    indexes: list[int] = []

    if source_mode in {"auto", "label"}:
        missing_labels = [value for value in values if _normalize_bin_label(value) not in label_to_index]
        if not missing_labels:
            indexes = [label_to_index[_normalize_bin_label(value)] for value in values]
        elif source_mode == "label":
            raise ValueError(f"bin group references unknown source_bin_labels={missing_labels}")

    if not indexes:
        try:
            indexes = [int(value) - 1 for value in values]
        except (TypeError, ValueError) as exc:
            raise ValueError(f"bin group values must match source labels or 1-based indexes, got {values}") from exc

    invalid_indexes = [index + 1 for index in indexes if index < 0 or index >= len(bins)]
    if invalid_indexes:
        raise ValueError(f"bin group references source_bin_indexes outside 1..{len(bins)}: {invalid_indexes}")
    return group_label, indexes


def _build_group_label_map(bins: list[dict[str, Any]], bin_groups: list[Any] | None) -> dict[str, Any]:
    if not bin_groups:
        return {_normalize_bin_label(bin_cfg.get("label")): bin_cfg.get("label") for bin_cfg in bins}

    label_map: dict[str, Any] = {}
    covered_indexes: list[int] = []
    for group_index, group_cfg in enumerate(bin_groups):
        configured_label, indexes = _resolve_group_indexes(group_cfg, bins)
        group_label = configured_label if configured_label is not None else group_index + 1
        for index in indexes:
            label_map[_normalize_bin_label(bins[index].get("label"))] = group_label
        covered_indexes.extend(indexes)

    expected_indexes = set(range(len(bins)))
    covered_set = set(covered_indexes)
    if covered_set != expected_indexes or len(covered_indexes) != len(covered_set):
        missing = sorted(expected_indexes - covered_set)
        duplicate = sorted(index for index in covered_indexes if covered_indexes.count(index) > 1)
        raise ValueError(
            "bin_groups must cover every configured source bin exactly once; "
            f"missing_source_bin_indexes={[i + 1 for i in missing]}, "
            f"duplicate_source_bin_indexes={[i + 1 for i in duplicate]}"
        )
    return label_map


def _apply_group_label_map(labels: pd.Series, label_map: dict[str, Any]) -> pd.Series:
    if not label_map:
        return labels
    return labels.map(lambda value: label_map.get(_normalize_bin_label(value), value) if not pd.isna(value) else value)


def apply_equal_frequency_binning(
    df: pd.DataFrame,
    score_field: str,
    bin_field: str,
    bin_count: int,
    bin_labels: list[Any],
    special_values: list[Any] | None,
    special_label_map: dict[float, Any],
    label_type: str | None,
) -> pd.DataFrame:
    """等频分箱：-1 等特殊值先单独打标，剩余正常分数再按样本量均分。"""
    out = df.copy()
    if score_field not in out.columns:
        out[bin_field] = pd.NA
        return out
    if bin_count <= 0:
        raise ValueError(f"bin_count must be positive, got {bin_count}")

    score = pd.to_numeric(out[score_field], errors="coerce")
    labels = pd.Series(pd.NA, index=out.index, dtype="object")
    special_mask = pd.Series(False, index=out.index)

    for value in _coerce_numeric_values(special_values):
        value_mask = score.eq(value)
        if value_mask.any():
            labels = labels.mask(value_mask, special_label_map.get(float(value), value))
            special_mask = special_mask | value_mask

    normal_score = score[score.notna() & ~special_mask]
    if not normal_score.empty:
        effective_count = min(bin_count, len(normal_score))
        # rank(method="first") 可在大量重复分数时仍保持各箱样本量尽量一致。
        ranked = normal_score.rank(method="first")
        assigned = pd.qcut(ranked, q=effective_count, labels=bin_labels[:effective_count])
        labels.loc[normal_score.index] = assigned.astype("object")

    out[bin_field] = _cast_bin_label(labels, label_type)
    return out


def apply_manual_threshold_binning(
    df: pd.DataFrame,
    score_field: str,
    bin_field: str,
    bins: list[dict[str, Any]],
    bin_groups: list[Any] | None,
    null_values: list[Any] | None,
    else_label: Any,
    label_type: str | None,
) -> pd.DataFrame:
    out = df.copy()
    if score_field not in out.columns:
        out[bin_field] = pd.NA
        return out
    if not bins:
        raise ValueError(f"Manual threshold binning requires bins: score_field={score_field}")

    score = pd.to_numeric(out[score_field], errors="coerce")
    labels = pd.Series(pd.NA, index=out.index, dtype="object")
    special_values = _coerce_numeric_values(null_values)
    special_mask = score.isin(special_values) if special_values else pd.Series(False, index=out.index)

    for value in special_values:
        labels = labels.mask(score.eq(value), value)

    normal_mask = score.notna() & ~special_mask
    last_index = len(bins) - 1
    for index, bin_cfg in enumerate(bins):
        label = bin_cfg.get("label")
        min_score = bin_cfg.get("min_score", float("-inf"))
        max_score = bin_cfg.get("max_score", float("inf"))
        if index == last_index:
            matched = normal_mask & score.ge(min_score) & score.le(max_score)
        else:
            matched = normal_mask & score.ge(min_score) & score.lt(max_score)
        labels = labels.mask(matched & labels.isna(), label)

    unmatched_mask = normal_mask & labels.isna()
    if else_label is not None:
        labels = labels.mask(unmatched_mask, else_label)

    labels = _apply_group_label_map(labels, _build_group_label_map(bins, bin_groups))
    out[bin_field] = _cast_bin_label(labels, label_type)
    return out


def _apply_one_score_binning(
    df: pd.DataFrame,
    score_cfg: dict[str, Any],
    binning_mode: str,
    scheme_name: str,
    logger,
) -> pd.DataFrame:
    score_field = score_cfg.get("score_field")
    bin_field = score_cfg.get("bin_field")
    if not score_field or not bin_field:
        logger.warning(f"Invalid score config skipped: {score_cfg}")
        return df

    label_type = score_cfg.get("bin_label_type")
    mode = str(score_cfg.get("binning_mode", binning_mode)).lower()
    if mode in {"quantile", "equal_frequency", "equal_freq"}:
        bin_count = int(score_cfg.get("bin_count", 5))
        out = apply_equal_frequency_binning(
            df,
            score_field=score_field,
            bin_field=bin_field,
            bin_count=bin_count,
            bin_labels=_quantile_bin_labels(score_cfg, label_type, bin_count),
            special_values=score_cfg.get("special_values", score_cfg.get("null_values", [-1])),
            special_label_map=_special_label_map(score_cfg),
            label_type=label_type,
        )
        effective_bin_cnt = bin_count
    elif mode in {"upper_bound", "manual", "threshold", "range"}:
        bins = score_cfg.get("bins", [])
        out = apply_manual_threshold_binning(
            df,
            score_field=score_field,
            bin_field=bin_field,
            bins=bins,
            bin_groups=score_cfg.get("bin_groups"),
            null_values=score_cfg.get("null_values", score_cfg.get("special_values", [-1])),
            else_label=score_cfg.get("else_label"),
            label_type=label_type,
        )
        effective_bin_cnt = len(score_cfg.get("bin_groups") or bins)
    else:
        raise ValueError(f"Unsupported binning_mode={mode} for score_field={score_field}")

    score = pd.to_numeric(out[score_field], errors="coerce") if score_field in out.columns else pd.Series(dtype=float)
    special_values = _coerce_numeric_values(score_cfg.get("null_values", score_cfg.get("special_values", [-1])))
    special_cnt = int(score.isin(special_values).sum()) if not score.empty else 0
    normal_cnt = int((score.notna() & ~score.isin(special_values)).sum()) if not score.empty else 0
    null_cnt = int(out[bin_field].isna().sum()) if bin_field in out.columns else len(out)
    logger.info(
        "Applied score binning: "
        f"scheme={scheme_name}, mode={mode}, model={score_cfg.get('name', score_field)}, "
        f"score_field={score_field}, bin_field={bin_field}, effective_bin_cnt={effective_bin_cnt:,}, "
        f"normal_score_cnt={normal_cnt:,}, special_score_cnt={special_cnt:,}, bin_null_cnt={null_cnt:,}"
    )
    return out


def apply_score_binning(df: pd.DataFrame, cfg: dict[str, Any], logger) -> pd.DataFrame:
    out = df.copy()
    for scheme_cfg in _score_binning_schemes(cfg):
        scheme_name = str(scheme_cfg.get("name", "default"))
        binning_mode = str(scheme_cfg.get("binning_mode", cfg.get("score_binning", {}).get("binning_mode", "equal_frequency")))
        for score_cfg in scheme_cfg.get("scores", []):
            out = _apply_one_score_binning(out, score_cfg, binning_mode, scheme_name, logger)
    return out


def add_sample_month(df: pd.DataFrame, cfg: dict[str, Any], logger) -> pd.DataFrame:
    month_col = cfg["analysis"]["sample_month_field"]
    candidates = [cfg["keys"]["datetime_key"], *cfg["keys"].get("datetime_aliases", [])]
    datetime_col = next((col for col in candidates if col in df.columns), None)
    if datetime_col is None:
        raise ValueError(f"None of the datetime columns exists: {candidates}")

    out = df.copy()
    parsed = pd.to_datetime(out[datetime_col], errors="coerce")
    out[month_col] = parsed.dt.strftime("%Y-%m")
    logger.info(f"Derived {month_col} from {datetime_col}, invalid_datetime_cnt={int(parsed.isna().sum()):,}")
    return out


def build_conversion_labels(df: pd.DataFrame, cfg: dict[str, Any], logger) -> pd.DataFrame:
    conv = cfg.get("conversion", {})
    out = df.copy()
    app_status_col = conv.get("application_status_field", "application_status")
    assess_col = conv.get("assessment_status_field", "assessment_status")
    deal_status_col = conv.get("deal_status_field", "status")
    fields = conv.get("output_fields", {})

    completed_col = fields.get("completed_flag", "is_completed_application")
    approved_col = fields.get("approved_flag", "is_approved_application")
    auto_col = fields.get("auto_approved_flag", "is_auto_approved_application")
    manual_col = fields.get("manual_approved_flag", "is_manual_approved_application")
    deal_col = fields.get("deal_flag", "is_deal_application")
    stage_col = fields.get("conversion_stage", "conversion_stage")

    for col in [app_status_col, assess_col, deal_status_col]:
        if col not in out.columns:
            logger.warning(f"Conversion source column missing: {col}; filling with NA")
            out[col] = pd.NA

    app_status = out[app_status_col].astype("string")
    assess_status = out[assess_col].astype("string")
    deal_status = out[deal_status_col].astype("string")

    incomplete_values = set(map(str, conv.get("incomplete_status_values", ["0.Incomplete", "1.In Progress"])))
    approved_prefixes = tuple(map(str, conv.get("approved_status_prefixes", ["3", "4"])))
    auto_keyword = str(conv.get("auto_approved_keyword", "Auto Approved"))
    manual_keyword = str(conv.get("manual_approved_keyword", "Manual Approved"))
    deal_values = set(map(str, conv.get("deal_status_values", [])))

    out[completed_col] = (~app_status.isin(incomplete_values) & app_status.notna()).astype(int)
    out[approved_col] = app_status.fillna("").str.slice(0, 1).isin(approved_prefixes).astype(int)
    out[auto_col] = ((out[approved_col] == 1) & assess_status.fillna("").str.contains(auto_keyword, case=False, regex=False)).astype(int)
    out[manual_col] = ((out[approved_col] == 1) & assess_status.fillna("").str.contains(manual_keyword, case=False, regex=False)).astype(int)
    out[deal_col] = deal_status.isin(deal_values).astype(int)

    conditions = [
        out[completed_col] == 0,
        (out[completed_col] == 1) & (out[approved_col] == 0),
        (out[approved_col] == 1) & (out[deal_col] == 0),
        (out[approved_col] == 1) & (out[deal_col] == 1),
    ]
    choices = ["incomplete_or_in_progress", "completed_not_approved", "approved_not_deal", "approved_and_deal"]
    out[stage_col] = np.select(conditions, choices, default="UNKNOWN")

    logger.info(
        "Built conversion labels: "
        f"completed_cnt={int(out[completed_col].sum()):,}, approved_cnt={int(out[approved_col].sum()):,}, "
        f"auto_approved_cnt={int(out[auto_col].sum()):,}, manual_approved_cnt={int(out[manual_col].sum()):,}, "
        f"deal_cnt={int(out[deal_col].sum()):,}"
    )
    return out


def apply_deal_amount_filter(df: pd.DataFrame, cfg: dict[str, Any], logger) -> pd.DataFrame:
    rules = cfg.get("amount_rules", {})
    total_col = rules.get("total_amount_field", "total_amount")
    out_col = rules.get("deal_total_amount_field", "deal_total_amount")
    status_col = rules.get("deal_status_field", "status")
    deal_values = set(map(str, rules.get("deal_status_values", [])))

    out = df.copy()
    if total_col not in out.columns or status_col not in out.columns:
        logger.warning(f"Amount/status field missing; {out_col} will be null")
        out[out_col] = np.nan
        return out

    amount = pd.to_numeric(out[total_col], errors="coerce")
    is_deal = out[status_col].astype("string").isin(deal_values)
    out[out_col] = amount.where(is_deal, np.nan)
    logger.info(
        f"Created deal amount field: {out_col}, valid_deal_amount_cnt={int(out[out_col].notna().sum()):,}, "
        f"deal_status_match_cnt={int(is_deal.sum()):,}"
    )
    return out


def build_analysis_dataset(cfg: dict[str, Any], logger: logging.Logger) -> pd.DataFrame:
    tables = load_input_tables(cfg, logger)
    enriched = enrich_base_sample(tables, cfg, logger)
    enriched = apply_score_binning(enriched, cfg, logger)
    enriched = add_sample_month(enriched, cfg, logger)
    enriched = build_conversion_labels(enriched, cfg, logger)
    return apply_deal_amount_filter(enriched, cfg, logger)


def safe_divide(numerator: float | int | None, denominator: float | int | None) -> float | None:
    if denominator is None or pd.isna(denominator) or denominator == 0:
        return None
    if numerator is None or pd.isna(numerator):
        return None
    return float(numerator) / float(denominator)


def _to_numeric_series(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(np.nan, index=df.index)
    return pd.to_numeric(df[col], errors="coerce")


def _to_category_series(df: pd.DataFrame, col: str, missing_label: str = "Missing") -> pd.Series:
    if col not in df.columns:
        return pd.Series(missing_label, index=df.index, dtype="object")
    values = df[col].astype("string").str.strip()
    return values.mask(values.isna() | values.eq(""), missing_label).astype("object")


def _calc_one_group(g: pd.DataFrame, total_rows: int, cfg: dict[str, Any], metric_groups: set[str] | None = None) -> dict[str, Any]:
    metric_groups = metric_groups or {"sample", "risk", "amount_risk", "mean", "conversion"}
    row: dict[str, Any] = {}

    if "sample" in metric_groups:
        row["sample_cnt"] = len(g)
        row["sample_pct"] = safe_divide(len(g), total_rows)

    if "risk" in metric_groups:
        for m in cfg.get("risk_metrics", {}).get("label_metrics", []):
            field = m.get("field")
            prefix = m.get("prefix", field)
            values = _to_numeric_series(g, field)
            bad_values = set(m.get("bad_values", [1]))
            good_values = set(m.get("good_values", [0]))
            bad_cnt = int(values.isin(bad_values).sum())
            good_cnt = int(values.isin(good_values).sum())
            valid_cnt = bad_cnt + good_cnt
            row[f"{prefix}_valid_cnt"] = valid_cnt
            row[f"{prefix}_bad_cnt"] = bad_cnt
            row[f"{prefix}_good_cnt"] = good_cnt
            row[f"{prefix}_bad_rate"] = safe_divide(bad_cnt, valid_cnt)

    if "amount_risk" in metric_groups:
        for m in cfg.get("risk_metrics", {}).get("amount_overdue_metrics", []):
            prefix = m.get("prefix", m.get("name"))
            dpd = _to_numeric_series(g, m.get("dpd_field"))
            numerator = _to_numeric_series(g, m.get("numerator_field"))
            denominator = _to_numeric_series(g, m.get("denominator_field"))
            valid_denominator = denominator.where(denominator > 0)
            overdue_mask = dpd.ge(float(m.get("overdue_threshold", 0))) & valid_denominator.notna()
            numerator_sum = float(numerator.where(overdue_mask, 0).fillna(0).sum())
            denominator_sum = float(valid_denominator.fillna(0).sum())
            row[f"{prefix}_overdue_amount"] = numerator_sum
            row[f"{prefix}_principal_amount"] = denominator_sum
            row[f"{prefix}_overdue_rate"] = safe_divide(numerator_sum, denominator_sum)

    if "mean" in metric_groups:
        for m in cfg.get("mean_metrics", []):
            src = m.get("source_field")
            out = m.get("output_field", f"avg_{src}")
            row[out] = float(_to_numeric_series(g, src).mean()) if src else None

    if "conversion" in metric_groups:
        fields = cfg.get("conversion", {}).get("output_fields", {})
        completed_col = fields.get("completed_flag", "is_completed_application")
        approved_col = fields.get("approved_flag", "is_approved_application")
        auto_col = fields.get("auto_approved_flag", "is_auto_approved_application")
        manual_col = fields.get("manual_approved_flag", "is_manual_approved_application")
        deal_col = fields.get("deal_flag", "is_deal_application")

        apply_cnt = len(g)
        completed_cnt = int(_to_numeric_series(g, completed_col).eq(1).sum())
        approved_cnt = int(_to_numeric_series(g, approved_col).eq(1).sum())
        auto_cnt = int(_to_numeric_series(g, auto_col).eq(1).sum())
        manual_cnt = int(_to_numeric_series(g, manual_col).eq(1).sum())
        deal_cnt = int(_to_numeric_series(g, deal_col).eq(1).sum())
        row.update(
            {
                "apply_cnt": apply_cnt,
                "completed_application_cnt": completed_cnt,
                "approved_application_cnt": approved_cnt,
                "auto_approved_application_cnt": auto_cnt,
                "manual_approved_application_cnt": manual_cnt,
                "deal_sample_cnt": deal_cnt,
                "completion_rate": safe_divide(completed_cnt, apply_cnt),
                "approval_rate": safe_divide(approved_cnt, completed_cnt),
                "auto_approval_rate": safe_divide(auto_cnt, completed_cnt),
                "manual_approval_rate": safe_divide(manual_cnt, completed_cnt),
                "auto_approval_share": safe_divide(auto_cnt, approved_cnt),
                "manual_approval_share": safe_divide(manual_cnt, approved_cnt),
                "deal_rate": safe_divide(deal_cnt, approved_cnt),
            }
        )

    return row


def calculate_group_metrics(
    df: pd.DataFrame,
    group_cols: list[str],
    cfg: dict[str, Any],
    include_metric_groups: list[str] | None = None,
) -> pd.DataFrame:
    missing_group_cols = [c for c in group_cols if c not in df.columns]
    if missing_group_cols:
        raise ValueError(f"Group columns missing: {missing_group_cols}")

    metric_groups = set(include_metric_groups) if include_metric_groups else None
    total_rows = len(df)
    records: list[dict[str, Any]] = []
    if not group_cols:
        records.append(_calc_one_group(df, total_rows, cfg, metric_groups))
        return pd.DataFrame(records)

    grouped = df.groupby(group_cols, dropna=False, sort=True)
    for keys, g in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        row = {col: val for col, val in zip(group_cols, keys)}
        row.update(_calc_one_group(g, total_rows, cfg, metric_groups))
        records.append(row)
    return pd.DataFrame(records)


def _numeric_cross_metric_configs(cfg: dict[str, Any]) -> list[dict[str, Any]]:
    return cfg.get("user_profile_metrics", {}).get("numeric_cross_metrics", [])


def _calc_one_user_profile_numeric_group(g: pd.DataFrame, cfg: dict[str, Any]) -> dict[str, Any]:
    row: dict[str, Any] = {"sample_cnt": len(g)}
    for m in _numeric_cross_metric_configs(cfg):
        agg = str(m.get("agg", "mean")).lower()
        src = m.get("source_field")
        out = m.get("output_field", f"avg_{src}")
        row[out] = float(_to_numeric_series(g, src).mean()) if src and agg == "mean" else None
    return row


def calculate_user_profile_numeric_metrics(df: pd.DataFrame, group_cols: list[str], cfg: dict[str, Any]) -> pd.DataFrame:
    missing_group_cols = [c for c in group_cols if c not in df.columns]
    if missing_group_cols:
        raise ValueError(f"Group columns missing: {missing_group_cols}")

    records: list[dict[str, Any]] = []
    if not group_cols:
        records.append(_calc_one_user_profile_numeric_group(df, cfg))
        return pd.DataFrame(records)

    grouped = df.groupby(group_cols, dropna=False, sort=True)
    for keys, g in grouped:
        if not isinstance(keys, tuple):
            keys = (keys,)
        row = {col: val for col, val in zip(group_cols, keys)}
        row.update(_calc_one_user_profile_numeric_group(g, cfg))
        records.append(row)
    return pd.DataFrame(records)


def _global_category_order(df: pd.DataFrame, field: str, top_n: int, missing_label: str, others_label: str) -> tuple[list[str], set[str]]:
    values = _to_category_series(df, field, missing_label)
    counts = values.value_counts(dropna=False)
    ranked = sorted(counts.items(), key=lambda item: (-int(item[1]), str(item[0])))
    top_values = [str(value) for value, _ in ranked[:top_n]]
    top_set = set(top_values)
    if len(ranked) > top_n:
        top_values.append(others_label)
    return top_values, top_set


def calculate_user_profile_category_distribution(
    df: pd.DataFrame,
    row_bin: str,
    metric_cfg: dict[str, Any],
    bin_values: list[Any],
) -> pd.DataFrame:
    if row_bin not in df.columns:
        raise ValueError(f"Group column missing: {row_bin}")

    field = metric_cfg.get("source_field")
    if not field or field not in df.columns:
        return pd.DataFrame()

    top_n = int(metric_cfg.get("top_n", 20))
    missing_label = metric_cfg.get("missing_label", "Missing")
    others_label = metric_cfg.get("others_label", "Others")
    category_order, top_set = _global_category_order(df, field, top_n, missing_label, others_label)

    working = df[[row_bin, field]].copy()
    values = _to_category_series(working, field, missing_label).astype(str)
    if others_label in category_order:
        values = values.where(values.isin(top_set), others_label)
    working["_category_value"] = values

    records: list[dict[str, Any]] = []
    for bin_value in bin_values:
        group = working[working[row_bin].astype(str) == str(bin_value)]
        sample_cnt = len(group)
        counts = group["_category_value"].value_counts(dropna=False).to_dict()
        for category_value in category_order:
            category_cnt = int(counts.get(category_value, 0))
            records.append(
                {
                    "profile_field": field,
                    "primary_model_score_bin": bin_value,
                    "category_value": category_value,
                    "sample_cnt": sample_cnt,
                    "category_cnt": category_cnt,
                    "category_pct": safe_divide(category_cnt, sample_cnt),
                }
            )
        records.append(
            {
                "profile_field": field,
                "primary_model_score_bin": bin_value,
                "category_value": "Total",
                "sample_cnt": sample_cnt,
                "category_cnt": sample_cnt,
                "category_pct": safe_divide(sample_cnt, sample_cnt),
            }
        )

    sample_cnt = len(working)
    counts = working["_category_value"].value_counts(dropna=False).to_dict()
    for category_value in category_order:
        category_cnt = int(counts.get(category_value, 0))
        records.append(
            {
                "profile_field": field,
                "primary_model_score_bin": "Total",
                "category_value": category_value,
                "sample_cnt": sample_cnt,
                "category_cnt": category_cnt,
                "category_pct": safe_divide(category_cnt, sample_cnt),
            }
        )
    records.append(
        {
            "profile_field": field,
            "primary_model_score_bin": "Total",
            "category_value": "Total",
            "sample_cnt": sample_cnt,
            "category_cnt": sample_cnt,
            "category_pct": safe_divide(sample_cnt, sample_cnt),
        }
    )
    return pd.DataFrame(records)


def _font(**kwargs) -> Font:
    return Font(name=EXCEL_FONT_NAME, **kwargs)


def _is_missing(value: Any) -> bool:
    return value is None or pd.isna(value)


def _label_key(value: Any) -> str:
    if _is_missing(value):
        return "__NA__"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _label_display(value: Any) -> str:
    if _is_missing(value):
        return "NA"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sort_values(values: list[Any]) -> list[Any]:
    def sort_key(value: Any) -> tuple[int, float, str]:
        if _is_missing(value):
            return (2, 0.0, "")
        try:
            return (0, float(value), str(value))
        except (TypeError, ValueError):
            return (1, 0.0, str(value))

    unique: dict[str, Any] = {}
    for value in values:
        unique.setdefault(_label_key(value), value)
    return sorted(unique.values(), key=sort_key)


def _configured_bin_order(cfg: dict[str, Any], bin_field: str, fallback: list[Any] | None = None) -> list[Any]:
    for scheme_cfg in _score_binning_schemes(cfg):
        for score_cfg in scheme_cfg.get("scores", []):
            if score_cfg.get("bin_field") != bin_field:
                continue
            labels: list[Any] = []
            special_values = score_cfg.get("special_values", score_cfg.get("null_values", []))
            labels.extend([score_cfg.get("special_label_map", {}).get(value, value) for value in special_values])

            mode = str(score_cfg.get("binning_mode", scheme_cfg.get("binning_mode", ""))).lower()
            if mode in {"quantile", "equal_frequency", "equal_freq"}:
                labels.extend(score_cfg.get("bin_labels", list(range(1, int(score_cfg.get("bin_count", 5)) + 1))))
            else:
                bins = score_cfg.get("bins", [])
                group_map = _build_group_label_map(bins, score_cfg.get("bin_groups"))
                labels.extend(group_map.values())
                else_label = score_cfg.get("else_label")
                if else_label is not None:
                    labels.append(group_map.get(_normalize_bin_label(else_label), else_label))
            return _sort_values(labels)
    return list(fallback or [1, 2, 3, 4, 5])


def _records_by_key(metrics: pd.DataFrame, group_cols: list[str]) -> dict[tuple[str, ...], pd.Series]:
    records: dict[tuple[str, ...], pd.Series] = {}
    for _, row in metrics.iterrows():
        records[tuple(_label_key(row[col]) for col in group_cols)] = row
    return records


def _cell_value(record: pd.Series | None, metric: str) -> Any:
    if record is None or metric not in record.index:
        return None
    value = record[metric]
    if _is_missing(value):
        return None
    if metric in COUNT_METRICS:
        return int(round(float(value)))
    return float(value) if isinstance(value, (int, float, np.integer, np.floating)) else value


def _metric_cell_format(metric: str) -> str:
    if metric.endswith("_pct") or metric.endswith("_rate") or metric.endswith("_share"):
        return "percent"
    if metric == "sample_cnt" or metric.endswith("_cnt"):
        return "integer"
    if metric.startswith("avg_"):
        return "decimal_2"
    return "value"


def _metric_number_format(metric: str) -> str | None:
    if metric in RATE_DENOMINATORS or _metric_cell_format(metric) == "percent":
        return "0.00%"
    if metric in COUNT_METRICS or _metric_cell_format(metric) == "integer":
        return "#,##0"
    if metric in AMOUNT_METRICS:
        return "#,##0.00"
    if metric in FOUR_DECIMAL_METRICS:
        return "0.0000"
    if _metric_cell_format(metric) == "decimal_2":
        return "#,##0.00"
    if _metric_cell_format(metric) == "value":
        return "0.0000"
    return None


def _user_profile_value(value: Any, metric: str) -> Any:
    if _is_missing(value) or value == "":
        return None
    cell_format = _metric_cell_format(metric)
    if cell_format == "integer":
        return int(round(float(value)))
    if cell_format in {"percent", "decimal_2", "value"}:
        return float(value)
    return value


def _write_dataframe(ws, df: pd.DataFrame, start_row: int) -> tuple[int, int]:
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            cell = ws.cell(row=start_row + row_offset, column=1 + col_offset, value=None if _is_missing(value) else value)
            cell.font = _font(color="333333")
    return len(rows), len(df.columns)


def _style_metric_title(cell) -> None:
    cell.font = _font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_table(ws, start_row: int, row_count: int, col_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = _font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="F3F6F8")
    total_corner_fill = PatternFill("solid", fgColor="E2F0D9")
    body_font = _font(color="333333")
    total_font = _font(bold=True, color="000000")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    end_row = start_row + row_count - 1

    for col in range(1, col_count + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(start_row + 1, end_row + 1):
        is_total_row = ws.cell(row=row, column=1).value == "Total"
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            is_total_col = ws.cell(row=start_row, column=col).value == "Total"
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = body_font
            if is_total_row or is_total_col:
                cell.fill = total_fill
                cell.font = total_font
            if is_total_row and is_total_col:
                cell.fill = total_corner_fill


def _apply_metric_number_format(ws, start_row: int, row_count: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        number_format = _metric_number_format(str(header))
        if not number_format:
            continue
        for row_idx in range(start_row + 1, start_row + row_count):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format


def _apply_pivot_number_format(ws, start_row: int, row_count: int, col_count: int, metric: str) -> None:
    number_format = _metric_number_format(metric)
    if not number_format:
        return
    for row in range(start_row + 1, start_row + row_count):
        for col in range(2, col_count + 1):
            ws.cell(row=row, column=col).number_format = number_format


def _auto_width(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for item in cell:
                if item.value is not None:
                    max_len = max(max_len, len(str(item.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _build_metric_pivot(df: pd.DataFrame, cfg: dict[str, Any], row_bin: str, column_bin: str, metric: str) -> pd.DataFrame:
    detail = calculate_group_metrics(df, [row_bin, column_bin], cfg)
    row_totals = calculate_group_metrics(df, [row_bin], cfg)
    column_totals = calculate_group_metrics(df, [column_bin], cfg)
    grand_total = calculate_group_metrics(df, [], cfg)

    row_values = _sort_values(list(detail[row_bin]) + list(row_totals[row_bin]))
    column_values = _sort_values(list(detail[column_bin]) + list(column_totals[column_bin]))
    detail_records = _records_by_key(detail, [row_bin, column_bin])
    row_total_records = _records_by_key(row_totals, [row_bin])
    column_total_records = _records_by_key(column_totals, [column_bin])
    grand_record = grand_total.iloc[0] if not grand_total.empty else None

    records: list[dict[str, Any]] = []
    for row_value in row_values:
        row_key = _label_key(row_value)
        record: dict[str, Any] = {row_bin: _label_display(row_value)}
        for column_value in column_values:
            column_key = _label_key(column_value)
            record[_label_display(column_value)] = _cell_value(detail_records.get((row_key, column_key)), metric)
        record["Total"] = _cell_value(row_total_records.get((row_key,)), metric)
        records.append(record)

    total_record: dict[str, Any] = {row_bin: "Total"}
    for column_value in column_values:
        column_key = _label_key(column_value)
        total_record[_label_display(column_value)] = _cell_value(column_total_records.get((column_key,)), metric)
    total_record["Total"] = _cell_value(grand_record, metric)
    records.append(total_record)
    return pd.DataFrame(records)


def _risk_perf_scheme_outputs(cfg: dict[str, Any], default_row_bin: str, default_column_bin: str) -> list[dict[str, Any]]:
    binning_cfg = cfg.get("score_binning", {})
    schemes_by_name = _score_binning_scheme_by_name(cfg)
    configured_names = binning_cfg.get("risk_perf_schemes")
    if configured_names:
        selected = [schemes_by_name[name] for name in configured_names if name in schemes_by_name]
    else:
        default_scheme = binning_cfg.get("default_scheme")
        selected = [schemes_by_name[default_scheme]] if default_scheme in schemes_by_name else _score_binning_schemes(cfg)[:1]

    outputs: list[dict[str, Any]] = []
    for index, scheme_cfg in enumerate(selected):
        scheme_name = str(scheme_cfg.get("name", f"scheme_{index + 1}"))
        outputs.append(
            {
                "title": scheme_cfg.get("title", scheme_name),
                "sheet_name": "risk_perf" if index == 0 else f"risk_perf_{scheme_name}",
                "row_bin": scheme_cfg.get("row_bin_field", default_row_bin),
                "column_bin": scheme_cfg.get("column_bin_field", default_column_bin),
            }
        )
    return outputs


def _write_risk_perf_sheet(
    wb: Workbook,
    sheet_name: str,
    title_prefix: str,
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    metrics: list[str],
) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    current_row = 1
    _style_metric_title(ws.cell(row=current_row, column=1, value=f"{title_prefix} - 全量申请样本交叉分布"))
    current_row += 1
    pivot = _build_metric_pivot(df, cfg, row_bin, column_bin, "sample_cnt")
    row_count, col_count = _write_dataframe(ws, pivot, current_row)
    _style_table(ws, current_row, row_count, col_count)
    _apply_pivot_number_format(ws, current_row, row_count, col_count, "sample_cnt")
    current_row += row_count + 2

    for metric in metrics:
        _style_metric_title(ws.cell(row=current_row, column=1, value=f"{title_prefix} - {metric}"))
        current_row += 1
        pivot = _build_metric_pivot(df, cfg, row_bin, column_bin, metric)
        row_count, col_count = _write_dataframe(ws, pivot, current_row)
        _style_table(ws, current_row, row_count, col_count)
        _apply_pivot_number_format(ws, current_row, row_count, col_count, metric)
        current_row += row_count + 2
    _auto_width(ws)


def _write_metric_group_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    metrics: list[str],
) -> None:
    ws = wb.create_sheet(sheet_name[:31])
    current_row = 1
    for metric in metrics:
        _style_metric_title(ws.cell(row=current_row, column=1, value=metric))
        current_row += 1
        pivot = _build_metric_pivot(df, cfg, row_bin, column_bin, metric)
        row_count, col_count = _write_dataframe(ws, pivot, current_row)
        _style_table(ws, current_row, row_count, col_count)
        _apply_pivot_number_format(ws, current_row, row_count, col_count, metric)
        current_row += row_count + 2
    _auto_width(ws)


def _deal_sample_df(df: pd.DataFrame, cfg: dict[str, Any], logger) -> pd.DataFrame:
    deal_col = cfg.get("conversion", {}).get("output_fields", {}).get("deal_flag", "is_deal_application")
    if deal_col not in df.columns:
        logger.warning(f"Deal sample flag missing; profile_deal_sample will be empty: {deal_col}")
        return df.iloc[0:0].copy()
    return df[pd.to_numeric(df[deal_col], errors="coerce").eq(1)].copy()


def export_risk_performance_workbook(
    output_path: Path,
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    logger,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "metric_guide"

    guide_df = pd.DataFrame(
        [
            {
                "category": category,
                "metric": metric,
                "cell_format": "percent" if metric in RATE_DENOMINATORS else "value",
                "denominator_metric": RATE_DENOMINATORS.get(metric, ""),
            }
            for category, metrics in METRIC_GROUPS.items()
            for metric in metrics
        ]
    )
    row_count, col_count = _write_dataframe(ws, guide_df, 1)
    _style_table(ws, 1, row_count, col_count)
    ws.freeze_panes = "A2"
    _auto_width(ws)

    for category, metrics in METRIC_GROUPS.items():
        if category == "risk_perf":
            for scheme_output in _risk_perf_scheme_outputs(cfg, row_bin, column_bin):
                _write_risk_perf_sheet(
                    wb=wb,
                    sheet_name=scheme_output["sheet_name"],
                    title_prefix=scheme_output["title"],
                    df=df,
                    cfg=cfg,
                    row_bin=scheme_output["row_bin"],
                    column_bin=scheme_output["column_bin"],
                    metrics=metrics,
                )
            continue

        _write_metric_group_sheet(wb, category, df, cfg, row_bin, column_bin, metrics)
        if category == "profile":
            _write_metric_group_sheet(
                wb,
                "profile_deal_sample",
                _deal_sample_df(df, cfg, logger),
                cfg,
                row_bin,
                column_bin,
                metrics,
            )

    raw_outputs = {
        "raw_cross_metrics": calculate_group_metrics(df, [row_bin, column_bin], cfg),
        "raw_row_totals": calculate_group_metrics(df, [row_bin], cfg),
        "raw_column_totals": calculate_group_metrics(df, [column_bin], cfg),
        "raw_grand_total": calculate_group_metrics(df, [], cfg),
    }
    for sheet_name, data in raw_outputs.items():
        ws = wb.create_sheet(sheet_name)
        row_count, col_count = _write_dataframe(ws, data, 1)
        _style_table(ws, 1, row_count, col_count)
        _apply_metric_number_format(ws, 1, row_count, list(data.columns))
        ws.freeze_panes = "C2" if sheet_name == "raw_cross_metrics" else "A2"
        _auto_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Wrote risk performance workbook: {output_path}")


def _user_profile_cfg_with_numeric_metrics(cfg: dict[str, Any], metrics: list[dict[str, Any]]) -> dict[str, Any]:
    user_profile_metrics = dict(cfg.get("user_profile_metrics", {}))
    user_profile_metrics["numeric_cross_metrics"] = metrics
    out = dict(cfg)
    out["user_profile_metrics"] = user_profile_metrics
    return out


def _valid_user_profile_numeric_metrics(df: pd.DataFrame, cfg: dict[str, Any], logger) -> list[dict[str, Any]]:
    valid: list[dict[str, Any]] = []
    for metric_cfg in cfg.get("user_profile_metrics", {}).get("numeric_cross_metrics", []):
        source_field = metric_cfg.get("source_field")
        if not source_field or source_field not in df.columns:
            logger.warning(f"User profile numeric field missing, skipped: {source_field}")
            continue
        if str(metric_cfg.get("agg", "mean")).lower() != "mean":
            logger.warning(f"Unsupported user profile numeric agg skipped: {metric_cfg}")
            continue
        valid.append(metric_cfg)
    return valid


def _valid_user_profile_category_metrics(df: pd.DataFrame, cfg: dict[str, Any], logger) -> list[dict[str, Any]]:
    valid: list[dict[str, Any]] = []
    for metric_cfg in cfg.get("user_profile_metrics", {}).get("category_distribution_metrics", []):
        source_field = metric_cfg.get("source_field")
        if not source_field or source_field not in df.columns:
            logger.warning(f"User profile category field missing, skipped: {source_field}")
            continue
        valid.append(metric_cfg)
    return valid


def _user_profile_numeric_metric_specs(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for metric_cfg in metrics:
        source_field = metric_cfg.get("source_field", "")
        metric = metric_cfg.get("output_field", f"avg_{source_field}")
        specs.append(
            {
                "category": metric_cfg.get("category", "user_profile"),
                "metric_type": "numeric_cross",
                "field_name": source_field,
                "metric_name": metric,
                "cell_format": _metric_cell_format(metric),
                "description": (
                    f"Mean value of {source_field} by primary_model_score_bin and comparison_model_score_bin."
                ),
            }
        )
    return specs


def _user_profile_category_metric_specs(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for metric_cfg in metrics:
        source_field = metric_cfg.get("source_field", "")
        sheet = metric_cfg.get("sheet", "category_distribution")
        top_n = int(metric_cfg.get("top_n", 3))
        specs.extend(
            [
                {
                    "category": sheet,
                    "metric_type": "category_distribution",
                    "field_name": source_field,
                    "metric_name": f"{source_field}_cnt",
                    "cell_format": "integer",
                    "description": f"Count distribution of {source_field} by primary_model_score_bin; keeps global Top{top_n} values and combines the rest into Others.",
                },
                {
                    "category": sheet,
                    "metric_type": "category_distribution",
                    "field_name": source_field,
                    "metric_name": f"{source_field}_pct",
                    "cell_format": "percent",
                    "description": f"Share distribution of {source_field} by primary_model_score_bin.",
                },
            ]
        )
    return specs


def _user_profile_metric_guide(numeric_metrics: list[dict[str, Any]], category_metrics: list[dict[str, Any]]) -> pd.DataFrame:
    columns = ["category", "metric_type", "field_name", "metric_name", "cell_format", "description"]
    records = _user_profile_numeric_metric_specs(numeric_metrics) + _user_profile_category_metric_specs(category_metrics)
    return pd.DataFrame(records, columns=columns)


def _user_profile_numeric_metric_groups(metrics: list[dict[str, Any]]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {}
    for metric_cfg in metrics:
        category = metric_cfg.get("category", "user_profile")
        metric = metric_cfg.get("output_field", f"avg_{metric_cfg.get('source_field', '')}")
        grouped.setdefault(category, []).append(metric)

    ordered: dict[str, list[str]] = {}
    for category in USER_PROFILE_CATEGORY_ORDER:
        if category in grouped:
            ordered[category] = grouped.pop(category)
    ordered.update(grouped)
    return ordered


def _category_metric_groups(metrics: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for metric_cfg in metrics:
        grouped.setdefault(metric_cfg.get("sheet", "category_distribution"), []).append(metric_cfg)

    ordered: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in USER_PROFILE_DISTRIBUTION_SHEET_ORDER:
        if sheet_name in grouped:
            ordered[sheet_name] = grouped.pop(sheet_name)
    ordered.update(grouped)
    return ordered


def _build_user_profile_numeric_pivot(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    row_bin: str,
    column_bin: str,
    metric: str,
    numeric_metrics: list[dict[str, Any]],
) -> pd.DataFrame:
    numeric_cfg = _user_profile_cfg_with_numeric_metrics(cfg, numeric_metrics)
    detail = calculate_user_profile_numeric_metrics(df, [row_bin, column_bin], numeric_cfg)
    row_totals = calculate_user_profile_numeric_metrics(df, [row_bin], numeric_cfg)
    column_totals = calculate_user_profile_numeric_metrics(df, [column_bin], numeric_cfg)
    grand_total = calculate_user_profile_numeric_metrics(df, [], numeric_cfg)
    row_order = _configured_bin_order(cfg, row_bin)
    column_order = _configured_bin_order(cfg, column_bin)

    detail_records = _records_by_key(detail, [row_bin, column_bin])
    row_total_records = _records_by_key(row_totals, [row_bin])
    column_total_records = _records_by_key(column_totals, [column_bin])
    grand_record = grand_total.iloc[0] if not grand_total.empty else None

    records: list[dict[str, Any]] = []
    for row_value in row_order:
        row_key = _label_key(row_value)
        record: dict[str, Any] = {row_bin: _label_display(row_value)}
        for column_value in column_order:
            cell_record = detail_records.get((row_key, _label_key(column_value)))
            record[_label_display(column_value)] = None if cell_record is None else _user_profile_value(cell_record.get(metric), metric)
        row_total_record = row_total_records.get((row_key,))
        record["Total"] = None if row_total_record is None else _user_profile_value(row_total_record.get(metric), metric)
        records.append(record)

    total_record: dict[str, Any] = {row_bin: "Total"}
    for column_value in column_order:
        column_total_record = column_total_records.get((_label_key(column_value),))
        total_record[_label_display(column_value)] = None if column_total_record is None else _user_profile_value(column_total_record.get(metric), metric)
    total_record["Total"] = None if grand_record is None else _user_profile_value(grand_record.get(metric), metric)
    records.append(total_record)
    return pd.DataFrame(records)


def _empty_user_profile_numeric_row(numeric_metrics: list[dict[str, Any]]) -> dict[str, Any]:
    row: dict[str, Any] = {"sample_cnt": 0}
    for metric_cfg in numeric_metrics:
        metric = metric_cfg.get("output_field", f"avg_{metric_cfg.get('source_field', '')}")
        row[metric] = None
    return row


def _complete_user_profile_numeric_metrics(
    metrics: pd.DataFrame,
    cfg: dict[str, Any],
    numeric_metrics: list[dict[str, Any]],
    group_cols: list[str],
) -> pd.DataFrame:
    if not group_cols:
        return metrics

    expected_values = {col: _configured_bin_order(cfg, col) for col in group_cols}
    existing = _records_by_key(metrics, group_cols) if not metrics.empty else {}
    records: list[dict[str, Any]] = []
    if len(group_cols) == 2:
        row_col, column_col = group_cols
        for row_value in expected_values[row_col]:
            for column_value in expected_values[column_col]:
                key = (_label_key(row_value), _label_key(column_value))
                if key in existing:
                    records.append(existing[key].to_dict())
                else:
                    row = {row_col: row_value, column_col: column_value}
                    row.update(_empty_user_profile_numeric_row(numeric_metrics))
                    records.append(row)
    else:
        return metrics

    columns = group_cols + ["sample_cnt"] + [
        metric_cfg.get("output_field", f"avg_{metric_cfg.get('source_field', '')}")
        for metric_cfg in numeric_metrics
    ]
    return pd.DataFrame(records).reindex(columns=columns)


def _build_category_distribution_pivot(distribution: pd.DataFrame, field: str, value_col: str, row_order: list[Any]) -> pd.DataFrame:
    field_data = distribution[distribution["profile_field"].eq(field)].copy()
    category_order = (
        field_data[
            field_data["primary_model_score_bin"].astype(str).eq("Total")
            & ~field_data["category_value"].astype(str).eq("Total")
        ]["category_value"]
        .astype(str)
        .tolist()
    )
    columns = category_order + ["Total"]
    records: list[dict[str, Any]] = []

    for row_value in row_order + ["Total"]:
        row_data = field_data[field_data["primary_model_score_bin"].astype(str).eq(str(row_value))]
        row_record: dict[str, Any] = {"primary_model_score_bin": _label_display(row_value)}
        for category_value in columns:
            match = row_data[row_data["category_value"].astype(str).eq(str(category_value))]
            if match.empty:
                row_record[category_value] = None
                continue
            value = match.iloc[0][value_col]
            metric_name = "category_pct" if value_col.endswith("_pct") else "category_cnt"
            row_record[category_value] = _user_profile_value(value, metric_name)
        records.append(row_record)

    return pd.DataFrame(records, columns=["primary_model_score_bin"] + columns)


def export_user_profile_workbook(
    df: pd.DataFrame,
    cfg: dict[str, Any],
    output_path: Path,
    row_bin: str,
    column_bin: str,
    logger,
) -> None:
    numeric_metrics = _valid_user_profile_numeric_metrics(df, cfg, logger)
    category_metrics = _valid_user_profile_category_metrics(df, cfg, logger)
    numeric_cfg = _user_profile_cfg_with_numeric_metrics(cfg, numeric_metrics)

    wb = Workbook()
    ws = wb.active
    ws.title = "metric_guide"
    guide_df = _user_profile_metric_guide(numeric_metrics, category_metrics)
    row_count, col_count = _write_dataframe(ws, guide_df, 1)
    _style_table(ws, 1, row_count, col_count)
    _apply_metric_number_format(ws, 1, row_count, list(guide_df.columns))
    ws.freeze_panes = "A2"
    _auto_width(ws)

    numeric_groups = _user_profile_numeric_metric_groups(numeric_metrics)
    category_groups = _category_metric_groups(category_metrics)

    def write_numeric_sheet(category: str) -> None:
        if category not in numeric_groups:
            return
        ws = wb.create_sheet(category)
        ws.freeze_panes = "A2"
        current_row = 1
        for metric in numeric_groups[category]:
            _style_metric_title(ws.cell(row=current_row, column=1, value=metric))
            current_row += 1
            pivot = _build_user_profile_numeric_pivot(df, cfg, row_bin, column_bin, metric, numeric_metrics)
            row_count, col_count = _write_dataframe(ws, pivot, current_row)
            _style_table(ws, current_row, row_count, col_count)
            _apply_pivot_number_format(ws, current_row, row_count, col_count, metric)
            current_row += row_count + 2
        _auto_width(ws)

    distribution_frames = [
        calculate_user_profile_category_distribution(df, row_bin, metric_cfg, _configured_bin_order(cfg, row_bin))
        for metric_cfg in category_metrics
    ]
    distribution = (
        pd.concat(distribution_frames, ignore_index=True)
        if distribution_frames
        else pd.DataFrame(columns=["profile_field", "primary_model_score_bin", "category_value", "sample_cnt", "category_cnt", "category_pct"])
    )

    def write_distribution_sheet(sheet_name: str) -> None:
        if sheet_name not in category_groups:
            return
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A2"
        current_row = 1
        for metric_cfg in category_groups[sheet_name]:
            field = metric_cfg["source_field"]
            for suffix, value_col in [("cnt", "category_cnt"), ("pct", "category_pct")]:
                metric_name = f"{field}_{suffix}"
                _style_metric_title(ws.cell(row=current_row, column=1, value=metric_name))
                current_row += 1
                pivot = _build_category_distribution_pivot(distribution, field, value_col, _configured_bin_order(cfg, row_bin))
                row_count, col_count = _write_dataframe(ws, pivot, current_row)
                _style_table(ws, current_row, row_count, col_count)
                _apply_pivot_number_format(ws, current_row, row_count, col_count, metric_name)
                current_row += row_count + 2
        _auto_width(ws)

    write_numeric_sheet("basic_profile")
    write_distribution_sheet("geo_profile_distribution")
    write_distribution_sheet("family_profile_distribution")
    write_distribution_sheet("attributed_category_distribution")
    write_numeric_sheet("occupation_profile")
    write_distribution_sheet("occupation_profile_distribution")

    raw_numeric = _complete_user_profile_numeric_metrics(
        calculate_user_profile_numeric_metrics(df, [row_bin, column_bin], numeric_cfg),
        cfg,
        numeric_metrics,
        [row_bin, column_bin],
    )
    raw_outputs = {
        "raw_user_profile_numeric_cross_metrics": raw_numeric,
        "raw_user_profile_category_distribution": distribution,
    }
    for sheet_name, data in raw_outputs.items():
        ws = wb.create_sheet(sheet_name)
        row_count, col_count = _write_dataframe(ws, data, 1)
        _style_table(ws, 1, row_count, col_count)
        _apply_metric_number_format(ws, 1, row_count, list(data.columns))
        ws.freeze_panes = "A2"
        _auto_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Wrote user profile workbook: {output_path}")


def _find_column_case_insensitive(columns: list[str], candidates: list[str]) -> str | None:
    lower_to_original = {str(col).strip().lower(): col for col in columns}
    for candidate in candidates:
        if candidate in lower_to_original:
            return lower_to_original[candidate]
    return None


def load_variable_library(path: Path, logger) -> tuple[pd.DataFrame, list[str]]:
    if not path.exists():
        raise FileNotFoundError(f"model_variable_library.csv not found: {path}")

    library_df = pd.read_csv(path, low_memory=False)
    if library_df.empty and not len(library_df.columns):
        raise ValueError(f"model_variable_library.csv has no columns: {path}")

    columns = [str(col) for col in library_df.columns]
    field_col = _find_column_case_insensitive(columns, VARIABLE_NAME_COLUMNS)
    if field_col:
        variables = (
            library_df[field_col]
            .dropna()
            .astype("string")
            .str.strip()
            .loc[lambda s: s.ne("")]
            .drop_duplicates()
            .tolist()
        )
    else:
        variables = [col for col in columns if col not in KEY_COLUMNS]
        logger.warning("No variable-name column found in model_variable_library.csv; treating wide-format non-key columns as variables.")

    if not variables:
        raise ValueError(f"No variables found in model_variable_library.csv: {path}")
    return library_df, variables


def merge_variable_library_data(
    df: pd.DataFrame,
    library_df: pd.DataFrame,
    variables: list[str],
    primary_key: str,
    logger,
) -> pd.DataFrame:
    if primary_key not in df.columns:
        raise ValueError(f"Main analysis data missing primary key: {primary_key}")

    out = df.copy()
    out[primary_key] = _normalize_key_series(out[primary_key])
    if primary_key not in library_df.columns:
        logger.warning(
            f"model_variable_library.csv does not contain {primary_key}; only variables already present in main analysis data can be analyzed."
        )
        return out

    merge_fields = [field for field in variables if field in library_df.columns and field not in out.columns]
    if not merge_fields:
        return out

    right = library_df[[primary_key] + merge_fields].copy()
    right[primary_key] = _normalize_key_series(right[primary_key])
    duplicate_rows = int(right.duplicated(subset=[primary_key], keep=False).sum())
    if duplicate_rows:
        logger.warning(f"model_variable_library.csv has duplicate application_id values; keeping first: duplicate_rows={duplicate_rows:,}")
        right = right.drop_duplicates(subset=[primary_key], keep="first")

    before_rows = len(out)
    out = out.merge(right, on=primary_key, how="left")
    if len(out) != before_rows:
        raise ValueError(f"Variable library merge changed row count: before={before_rows}, after={len(out)}")
    logger.info(f"Merged model variable library fields: added_field_cnt={len(merge_fields):,}")
    return out


def _existing_and_missing_fields(df: pd.DataFrame, variables: list[str]) -> tuple[list[str], list[str]]:
    existing = [field for field in variables if field in df.columns]
    missing = [field for field in variables if field not in df.columns]
    return existing, missing


def _mean_or_none(series: pd.Series) -> float | None:
    non_null = series.dropna()
    if non_null.empty:
        return None
    return float(non_null.mean())


def build_mean_by_bin(df: pd.DataFrame, variables: list[str], bin_col: str) -> pd.DataFrame:
    if bin_col not in df.columns:
        raise ValueError(f"Main analysis data missing {bin_col}")

    records: list[dict[str, Any]] = []
    for bin_value in _sort_values(df[bin_col].drop_duplicates().tolist()):
        group = df[df[bin_col].astype(str).eq(str(bin_value))]
        total_cnt = len(group)
        for variable in variables:
            numeric = pd.to_numeric(group[variable], errors="coerce")
            non_null_cnt = int(numeric.notna().sum())
            records.append(
                {
                    bin_col: bin_value,
                    "variable_name": variable,
                    "avg_value": float(numeric.mean()) if non_null_cnt else None,
                    "non_null_cnt": non_null_cnt,
                    "total_cnt": total_cnt,
                    "missing_pct": safe_divide(total_cnt - non_null_cnt, total_cnt),
                }
            )
    return pd.DataFrame(records)


def build_summary_by_variable(df: pd.DataFrame, variables: list[str]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    total_cnt = len(df)
    for variable in variables:
        numeric = pd.to_numeric(df[variable], errors="coerce")
        non_null_cnt = int(numeric.notna().sum())
        missing_cnt = total_cnt - non_null_cnt
        records.append(
            {
                "variable_name": variable,
                "avg_value": float(numeric.mean()) if non_null_cnt else None,
                "non_null_cnt": non_null_cnt,
                "total_cnt": total_cnt,
                "missing_cnt": missing_cnt,
                "missing_pct": safe_divide(missing_cnt, total_cnt),
                "min_value": float(numeric.min()) if non_null_cnt else None,
                "max_value": float(numeric.max()) if non_null_cnt else None,
            }
        )
    return pd.DataFrame(records)


def build_cross_bin_mean_pivot(df: pd.DataFrame, variable: str, primary_bin_col: str, comparison_bin_col: str) -> pd.DataFrame:
    missing_cols = [col for col in [primary_bin_col, comparison_bin_col] if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Main analysis data missing columns: {missing_cols}")

    primary_values = _sort_values(df[primary_bin_col].drop_duplicates().tolist())
    comparison_values = _sort_values(df[comparison_bin_col].drop_duplicates().tolist())
    numeric = pd.to_numeric(df[variable], errors="coerce")

    records: list[dict[str, Any]] = []
    for primary_value in primary_values:
        row_mask = df[primary_bin_col].astype(str).eq(str(primary_value))
        record: dict[str, Any] = {primary_bin_col: primary_value}
        for comparison_value in comparison_values:
            cell_mask = row_mask & df[comparison_bin_col].astype(str).eq(str(comparison_value))
            record[str(comparison_value)] = _mean_or_none(numeric[cell_mask])
        record["Total"] = _mean_or_none(numeric[row_mask])
        records.append(record)

    total_record: dict[str, Any] = {primary_bin_col: "Total"}
    for comparison_value in comparison_values:
        column_mask = df[comparison_bin_col].astype(str).eq(str(comparison_value))
        total_record[str(comparison_value)] = _mean_or_none(numeric[column_mask])
    total_record["Total"] = _mean_or_none(numeric)
    records.append(total_record)
    return pd.DataFrame(records, columns=[primary_bin_col] + [str(v) for v in comparison_values] + ["Total"])


def _style_feature_flat_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    headers = [cell.value for cell in ws[1]]
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = _font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = headers[cell.column - 1]
            cell.border = border
            if header in {"avg_value", "min_value", "max_value"}:
                cell.number_format = "0.0000"
            elif header == "missing_pct":
                cell.number_format = "0.00%"
            elif header in {"non_null_cnt", "total_cnt", "missing_cnt"}:
                cell.number_format = "#,##0"


def _safe_sheet_name(base_name: str, part: int | None = None) -> str:
    suffix = "" if part is None else f"_{part}"
    return f"{base_name[:31 - len(suffix)]}{suffix}"


def _write_split_dataframe(wb: Workbook, base_sheet_name: str, df: pd.DataFrame) -> None:
    max_data_rows = MAX_EXCEL_ROWS - 1
    if len(df) <= max_data_rows:
        ws = wb.create_sheet(_safe_sheet_name(base_sheet_name))
        _write_dataframe(ws, df, 1)
        _style_feature_flat_sheet(ws)
        _auto_width(ws)
        return

    start = 0
    part = 1
    while start < len(df):
        chunk = df.iloc[start : start + max_data_rows]
        ws = wb.create_sheet(_safe_sheet_name(base_sheet_name, part))
        _write_dataframe(ws, chunk, 1)
        _style_feature_flat_sheet(ws)
        _auto_width(ws)
        start += max_data_rows
        part += 1


def _write_cross_bin_mean_pivot_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    variables: list[str],
    primary_bin_col: str,
    comparison_bin_col: str,
) -> None:
    ws = wb.create_sheet("cross_bin_distribution")
    ws.freeze_panes = "A2"
    current_row = 1
    for variable in variables:
        _style_metric_title(ws.cell(row=current_row, column=1, value=f"{variable}_avg"))
        current_row += 1
        pivot = build_cross_bin_mean_pivot(df, variable, primary_bin_col, comparison_bin_col)
        row_count, col_count = _write_dataframe(ws, pivot, current_row)
        _style_table(ws, current_row, row_count, col_count)
        for row in range(current_row + 1, current_row + row_count):
            for col in range(2, col_count + 1):
                ws.cell(row=row, column=col).number_format = "0.0000"
        current_row += row_count + 2
    _auto_width(ws)


def export_feature_mean_profile_workbook(
    output_path: Path,
    df: pd.DataFrame,
    cfg: dict[str, Any],
    primary_bin_col: str,
    comparison_bin_col: str,
    logger,
) -> None:
    library_df, variables = load_variable_library(Path(cfg["feature_profile"]["variable_library_path"]), logger)
    analysis_df = merge_variable_library_data(df, library_df, variables, cfg["keys"]["primary_key"], logger)
    if primary_bin_col not in analysis_df.columns:
        raise ValueError(f"Main analysis data missing {primary_bin_col}")
    if comparison_bin_col not in analysis_df.columns:
        raise ValueError(f"Main analysis data missing {comparison_bin_col}")

    existing_fields, missing_field_names = _existing_and_missing_fields(analysis_df, variables)
    if not existing_fields:
        raise ValueError("No variables from model_variable_library.csv matched main analysis data columns.")

    summary = build_summary_by_variable(analysis_df, existing_fields)
    all_null_cnt = int(summary["non_null_cnt"].eq(0).sum())
    for variable in summary.loc[summary["non_null_cnt"].eq(0), "variable_name"]:
        logger.warning(f"Variable has no numeric values after pd.to_numeric conversion: {variable}")

    logger.info(f"Variable list read count: {len(variables):,}")
    logger.info(f"Variables matched count: {len(existing_fields):,}")
    logger.info(f"Variables missing count: {len(missing_field_names):,}")
    logger.info(f"All-null or non-numeric variable count: {all_null_cnt:,}")

    primary_distribution = build_mean_by_bin(analysis_df, existing_fields, primary_bin_col)
    comparison_distribution = build_mean_by_bin(analysis_df, existing_fields, comparison_bin_col)

    wb = Workbook()
    wb.remove(wb.active)
    _write_split_dataframe(wb, "primary_bin_distribution", primary_distribution)
    _write_split_dataframe(wb, "comparison_bin_distribution", comparison_distribution)
    _write_cross_bin_mean_pivot_sheet(wb, analysis_df, existing_fields, primary_bin_col, comparison_bin_col)
    _write_split_dataframe(wb, "summary_by_variable", summary)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Wrote feature mean profile workbook: {output_path}")
